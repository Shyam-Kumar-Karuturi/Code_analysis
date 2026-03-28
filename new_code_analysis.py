import os
import time
import random
import pandas as pd
import numpy as np
import google.generativeai as genai
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from dotenv import load_dotenv

# ---------------------------
# CONFIGURATION
# ---------------------------
load_dotenv()
API_KEY = os.getenv("GOOGLE_API_KEY")

# Fallback Key
if not API_KEY:
    API_KEY = ''

if not API_KEY:
    print("CRITICAL WARNING: No Google API Key found. Script will fail.")

genai.configure(api_key=API_KEY)
EMBED_MODEL = "models/text-embedding-004"

# ---------------------------
# GLOBAL CACHE (The Speed Booster)
# ---------------------------
# This dictionary stores embeddings for text we've already seen.
# It persists across all states during the run.
_embedding_cache = {}


# ---------------------------
# Helpers
# ---------------------------
def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Strip whitespace/newlines from column names."""
    df.columns = df.columns.str.strip().str.replace('\n', ' ').str.replace('\r', ' ')
    return df


def find_column(df: pd.DataFrame, possible_names):
    """Find column name in df that matches one of possible_names (case-insensitive)."""
    lower_targets = [p.lower() for p in possible_names]
    for col in df.columns:
        if col.lower().strip() in lower_targets:
            return col
    return None


def retry_with_backoff(retries=3, backoff_in_seconds=1):
    """Decorator to retry API calls on failure."""

    def decorator(func):
        def wrapper(*args, **kwargs):
            x = 0
            while True:
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    if x == retries:
                        print(f"    Failed after {retries} retries: {e}")
                        raise e
                    sleep = (backoff_in_seconds * 2 ** x) + random.uniform(0, 1)
                    time.sleep(sleep)
                    x += 1

        return wrapper

    return decorator


@retry_with_backoff(retries=5, backoff_in_seconds=2)
def generate_embedding_from_api(text: str):
    """Actual API call to Google Gemini."""
    # Rate limit prevention (small sleep only on actual API calls)
    time.sleep(0.4)
    resp = genai.embed_content(
        model=EMBED_MODEL,
        content=str(text),
        task_type="semantic_similarity"
    )
    return np.array(resp["embedding"])


def embed(text: str):
    """
    Wrapper that checks CACHE first before calling the API.
    """
    if text is None or str(text).strip() == "":
        text = "empty"

    text_str = str(text).strip()

    # 1. CHECK CACHE
    if text_str in _embedding_cache:
        return _embedding_cache[text_str]

    # 2. CALL API (If not in cache)
    try:
        vector = generate_embedding_from_api(text_str)
        # 3. SAVE TO CACHE
        _embedding_cache[text_str] = vector
        return vector
    except Exception:
        # Return zero vector on total failure to allow script to continue
        return np.zeros(768)


def cosine_similarity(v1, v2):
    """Cosine similarity between two numpy vectors."""
    n1 = np.linalg.norm(v1)
    n2 = np.linalg.norm(v2)
    if n1 == 0 or n2 == 0:
        return 0.0
    return float(np.dot(v1, v2) / (n1 * n2))


def write_summary(ws, summary: dict):
    ws.insert_rows(1, amount=10)
    ws["A1"] = "SUMMARY"
    ws["A1"].font = Font(bold=True)
    row = 2
    for key, val in summary.items():
        ws[f"A{row}"] = key
        ws[f"B{row}"] = val
        row += 1


def apply_conditional_formatting(ws):
    colors = {
        "Severe Change": "FFC7CE",  # Red
        "Moderate Change": "FFEB9C",  # Yellow
        "Minor Wording Change": "C6EFCE",  # Green
        "New Entry": "BDD7EE",  # Blue
        "No Change": "FFFFFF",  # White
        "Medicaid Change": "E1D3F6"  # Purple-ish
    }

    header_row = 11
    severity_col_idx = 7  # Default

    # Find Severity column
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val and "Severity" in str(val):
            severity_col_idx = col
            break

    for row in range(header_row + 1, ws.max_row + 1):
        cell = ws.cell(row=row, column=severity_col_idx)
        severity_val = str(cell.value) if cell.value is not None else ""
        fill_color = None

        if "Medicaid Change" in severity_val:
            fill_color = colors["Medicaid Change"]
        elif "Severe Change" in severity_val:
            fill_color = colors["Severe Change"]
        elif "Moderate Change" in severity_val:
            fill_color = colors["Moderate Change"]
        elif "Minor Wording Change" in severity_val:
            fill_color = colors["Minor Wording Change"]
        elif "New Entry" in severity_val:
            fill_color = colors["New Entry"]
        elif "No Change" in severity_val:
            fill_color = colors["No Change"]

        if fill_color:
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill


# ---------------------------
# CORE LOGIC
# ---------------------------
def compare_notes_and_medicaid(q3_df: pd.DataFrame, q4_df: pd.DataFrame, notes_col: str, medicaid_col: str = None):
    q3_df = q3_df.copy()
    q4_df = q4_df.copy()

    # Normalize 'Code' column
    code_col_q3 = find_column(q3_df, ["Code", "CPT Code", "Procedure Code", "Service Code", "HCPCS"])
    code_col_q4 = find_column(q4_df, ["Code", "CPT Code", "Procedure Code", "Service Code", "HCPCS"])

    if not code_col_q3 or not code_col_q4: return None

    # Standardize Codes
    q3_df["Code"] = q3_df[code_col_q3].astype(str).str.strip()
    q4_df["Code"] = q4_df[code_col_q4].astype(str).str.strip()

    # --- CRITICAL FIX: DEDUPLICATE CODES ---
    q3_df = q3_df.drop_duplicates(subset=["Code"], keep="first")
    q4_df = q4_df.drop_duplicates(subset=["Code"], keep="first")

    # Create lookup dictionary
    q4_lookup = q4_df.set_index("Code").to_dict('index')

    report_rows = []
    has_med_col = medicaid_col is not None

    # 1. Process OLD codes
    for _, row in q3_df.iterrows():
        code = row["Code"]
        q3_notes = str(row.get(notes_col)).strip() if notes_col else ""
        if q3_notes == "nan" or q3_notes == "None": q3_notes = ""

        q3_med = str(row.get(medicaid_col)).strip() if has_med_col else ""
        if q3_med == "nan" or q3_med == "None": q3_med = ""

        q3_val_str = f"Notes: {q3_notes}"
        if has_med_col: q3_val_str += f" | Medicaid: {q3_med}"

        if code not in q4_lookup:
            report_rows.append({
                "Code": code, "Status": "Removed in 2026 Q1",
                "Severity": "Severe Change",
                "2025 Q4 Value": q3_val_str, "2026 Q1 Value": ""
            })
            continue

        row_new = q4_lookup[code]
        q4_notes = str(row_new.get(notes_col)).strip() if notes_col else ""
        if q4_notes == "nan" or q4_notes == "None": q4_notes = ""

        q4_med = str(row_new.get(medicaid_col)).strip() if has_med_col else ""
        if q4_med == "nan" or q4_med == "None": q4_med = ""

        q4_val_str = f"Notes: {q4_notes}"
        if has_med_col: q4_val_str += f" | Medicaid: {q4_med}"

        notes_same = (q3_notes == q4_notes)
        med_same = (q3_med == q4_med) if has_med_col else True

        if notes_same and med_same:
            report_rows.append({
                "Code": code, "Status": "No Change",
                "Severity": "No Change",
                "2025 Q4 Value": q3_val_str, "2026 Q1 Value": q4_val_str
            })
        else:
            severity_parts = []
            if has_med_col and not med_same: severity_parts.append("Medicaid Change")

            if not notes_same:
                sim = cosine_similarity(embed(q3_notes), embed(q4_notes))
                if sim < 0.6:
                    severity_parts.append("Severe Change")
                elif sim < 0.85:
                    severity_parts.append("Moderate Change")
                else:
                    severity_parts.append("Minor Wording Change")

            report_rows.append({
                "Code": code, "Status": "Modified",
                "Severity": "; ".join(severity_parts),
                "2025 Q4 Value": q3_val_str, "2026 Q1 Value": q4_val_str
            })

    # 2. Process NEW codes
    q3_codes = set(q3_df["Code"])
    new_codes = set(q4_lookup.keys()) - q3_codes

    for code in new_codes:
        row_new = q4_lookup[code]
        q4_notes = str(row_new.get(notes_col)).strip() if notes_col else ""
        if q4_notes == "nan" or q4_notes == "None": q4_notes = ""

        q4_med = str(row_new.get(medicaid_col)).strip() if has_med_col else ""
        if q4_med == "nan" or q4_med == "None": q4_med = ""

        q4_val_str = f"Notes: {q4_notes}"
        if has_med_col: q4_val_str += f" | Medicaid: {q4_med}"

        report_rows.append({
            "Code": code, "Status": "New in 2026 Q1",
            "Severity": "New Entry",
            "2025 Q4 Value": "", "2026 Q1 Value": q4_val_str
        })

    return pd.DataFrame(report_rows, columns=["Code", "Status", "Severity", "2025 Q4 Value", "2026 Q1 Value"])
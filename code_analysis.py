import os
import pandas as pd
import numpy as np
import google.generativeai as genai
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from dotenv import load_dotenv
load_dotenv()

genai.configure(api_key=os.getenv("GOOGLE_API_KEY"))
EMBED_MODEL = "models/text-embedding-004"

def normalize_columns(df):
    df.columns = df.columns.str.strip().str.replace('\n', ' ').str.replace('\r', ' ')
    return df

def find_column(df, possible_names):
    for col in df.columns:
        if col.lower().strip() in [name.lower() for name in possible_names]:
            return col
    raise KeyError(f"❌ None of these columns found: {possible_names}")

def embed(text: str):
    """Generate embedding vector using Gemini embeddings."""
    if text is None or str(text).strip() == "":
        text = "empty"
    emb = genai.embed_content(
        model=EMBED_MODEL,
        content=str(text),
        task_type="semantic_similarity"
    )
    return np.array(emb["embedding"])


def cosine_similarity(v1, v2):
    return float(np.dot(v1, v2) / (np.linalg.norm(v1) * np.linalg.norm(v2)))


def compute_summary(df):
    """Creates a summary dict for difference summary section."""
    return {
        "Total in Q3": len(df),
        "No Change": (df["Status"] == "No Change").sum(),
        "Modified": (df["Status"] == "Modified").sum(),
        "Severe Change": (df["Severity"] == "Severe Change").sum(),
        "Moderate Change": (df["Severity"] == "Moderate Change").sum(),
        "Minor Change": (df["Severity"] == "Minor Wording Change").sum(),
        "New in Q4": (df["Status"] == "New in Q4").sum(),
        "Removed in Q4": (df["Status"] == "Removed in Q4").sum()
    }


def semantic_compare(q3_df, q4_df, text_column):
    q3_df["Code"] = q3_df["Code"].astype(str)
    q4_df["Code"] = q4_df["Code"].astype(str)

    q4_lookup = {row["Code"]: row for _, row in q4_df.iterrows()}
    report_rows = []

    for _, q3_row in q3_df.iterrows():
        code = q3_row["Code"]

        if code not in q4_lookup:
            report_rows.append({
                "Code": code,
                "Status": "Removed in Q4",
                "Column": text_column,
                "Q3 Value": q3_row[text_column],
                "Q4 Value": "",
                "Similarity": "",
                "Severity": "Severe Change"
            })
            continue

        q4_row = q4_lookup[code]

        old = str(q3_row[text_column]).strip()
        new = str(q4_row[text_column]).strip()

        if old == new:
            report_rows.append({
                "Code": code,
                "Status": "No Change",
                "Column": text_column,
                "Q3 Value": old,
                "Q4 Value": new,
                "Similarity": 1.0,
                "Severity": "No Change"
            })
            continue

        old_emb = embed(old)
        new_emb = embed(new)
        sim = cosine_similarity(old_emb, new_emb)

        severity = (
            "Severe Change" if sim < 0.55 else
            "Moderate Change" if sim < 0.80 else
            "Minor Wording Change"
        )

        report_rows.append({
            "Code": code,
            "Status": "Modified",
            "Column": text_column,
            "Q3 Value": old,
            "Q4 Value": new,
            "Similarity": round(sim, 4),
            "Severity": severity
        })

    # New in Q4
    q3_codes = set(q3_df["Code"])
    q4_codes = set(q4_df["Code"])
    new_codes = q4_codes - q3_codes

    for code in new_codes:
        row = q4_df[q4_df["Code"] == code].iloc[0]
        report_rows.append({
            "Code": code,
            "Status": "New in Q4",
            "Column": text_column,
            "Q3 Value": "",
            "Q4 Value": row[text_column],
            "Similarity": "",
            "Severity": "New Entry"
        })

    return pd.DataFrame(report_rows)


def apply_conditional_formatting(ws):
    """Apply color formatting based on Severity values."""
    colors = {
        "Severe Change": "FFC7CE",     # Red
        "Moderate Change": "FFEB9C",   # Yellow
        "Minor Wording Change": "C6EFCE", # Green
        "New Entry": "BDD7EE",         # Blue
        "No Change": "FFFFFF"          # White
    }

    for row in range(12, ws.max_row + 1):  # Start after summary
        severity = ws[f"G{row}"].value  # G column = Severity

        if severity in colors:
            fill = PatternFill(start_color=colors[severity], end_color=colors[severity], fill_type="solid")
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill


def write_summary(ws, summary):
    """Insert summary rows at top of sheet."""
    ws.insert_rows(1, amount=10)

    ws["A1"] = "SUMMARY"
    ws["A1"].font = Font(bold=True)

    row = 2
    for key, val in summary.items():
        ws[f"A{row}"] = key
        ws[f"B{row}"] = val
        row += 1


def process_file(excel_path):
    print("Reading Excel...")
    xls = pd.ExcelFile(excel_path)

    wa_q3 = normalize_columns(pd.read_excel(xls, "WA Q3"))
    wa_q4 = normalize_columns(pd.read_excel(xls, "WA Q4"))
    md_q3 = normalize_columns(pd.read_excel(xls, "Medicaid Q3"))
    md_q4 = normalize_columns(pd.read_excel(xls, "Medicaid Q4"))

    wa_q3 = wa_q3.replace(np.nan, "", regex=True)
    wa_q4 = wa_q4.replace(np.nan, "", regex=True)
    md_q3 = md_q3.replace(np.nan, "", regex=True)
    md_q4 = md_q4.replace(np.nan, "", regex=True)

    # Auto-detect column names
    wa_col_q3 = find_column(wa_q3, ["Code Notes"])
    wa_col_q4 = find_column(wa_q4, ["Code Notes"])

    print("Comparing WA...")
    wa_report = semantic_compare(wa_q3, wa_q4, wa_col_q3)

    md_col_q3 = find_column(md_q3, ["MHI Code Notes"])
    md_col_q4 = find_column(md_q4, ["MHI Code Notes"])

    print("Comparing Medicaid...")
    md_report = semantic_compare(md_q3, md_q4, md_col_q3)

    # Replace NaN with empty strings before exporting
    wa_report = wa_report.replace(np.nan, "", regex=True)
    md_report = md_report.replace(np.nan, "", regex=True)

    with pd.ExcelWriter(excel_path, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer:
        wa_report.to_excel(writer, sheet_name="WA Q3 vs WA Q4", index=False)
        md_report.to_excel(writer, sheet_name="Medicaid Q3 vs Medicaid Q4", index=False)

    wb = load_workbook(excel_path)

    # Apply summary + formatting to WA sheet
    ws_wa = wb["WA Q3 vs WA Q4"]
    summary_wa = compute_summary(wa_report)
    write_summary(ws_wa, summary_wa)
    apply_conditional_formatting(ws_wa)

    # Apply summary + formatting to Medicaid sheet
    ws_md = wb["Medicaid Q3 vs Medicaid Q4"]
    summary_md = compute_summary(md_report)
    write_summary(ws_md, summary_md)
    apply_conditional_formatting(ws_md)

    wb.save(excel_path)

    print("\n🎉 Done!")
    print(f"Sheets updated in: {excel_path}")
    print("→ WA Q3 vs WA Q4")
    print("→ Medicaid Q3 vs Medicaid Q4")
    print("✔ Color-coded")
    print("✔ Includes summary section")


if __name__ == "__main__":
    process_file("Authorization Business Matrix 2025 Q3 - WA and Medicaid - Reference.xlsx")

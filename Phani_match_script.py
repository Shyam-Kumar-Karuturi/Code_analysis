import ast
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ─────────────────────────────────────────────
# CONSTANTS  –  edit these before running
# ─────────────────────────────────────────────

# List the sheet names you want the script to process.
# Leave empty  →  process ALL sheets in the workbook.
TARGET_SHEETS: list[str] = [
    "MS",
    "ID",
    "NY",
]

# Value to place in every data cell of the "api response" column.
# Allowed values:  ""  (empty)  |  "Match"  |  "No Match"
MATCH_VALUE: str = ""

# Input file path
INPUT_FILE: str = "/Users/xaiuser/Downloads/Copy of batch_report_UNV_367_2026-03.xlsx"

# Directory where per-sheet output files are saved (use "." for current folder)
OUTPUT_DIR: str = "."

# Maximum allowed count for DENY / OKAY statuses from recommendation column
STATUS_MAX: int = 15

# ─────────────────────────────────────────────
# Internal column name constants
# ─────────────────────────────────────────────
API_RESPONSE_COL  = "api response"
JUSTIFICATION_COL = "justification"
RECOMMENDATION_COL = "recommendation"
YELLOW_HEX = "FFFF00"

# ─────────────────────────────────────────────


def parse_recommendation(value) -> dict | None:
    """
    Parse a single cell value from the recommendation column.
    The cell contains a Python-dict-like string, e.g.:
      {'recommendation': [{'actionType': '...', 'status': 'DENY', ...}], ...}
    Returns the parsed dict, or None if parsing fails.
    """
    if pd.isna(value) or not str(value).strip():
        return None
    try:
        return ast.literal_eval(str(value))
    except Exception:
        return None


def get_output_filename(sheet_name: str, df: pd.DataFrame) -> str:
    """
    Determine the output filename for a sheet:
      - If the 'justification' column contains ANY value starting with 'Marketplace'
          →  {SHEET_NAME}_Validation_Marketplace.xlsx
      - Otherwise
          →  {SHEET_NAME}_Validation_Medicaid.xlsx
    """
    label = "Medicaid"
    if JUSTIFICATION_COL in df.columns:
        starts_with_marketplace = (
            df[JUSTIFICATION_COL]
            .astype(str)
            .str.startswith("Marketplace", na=False)
            .any()
        )
        if starts_with_marketplace:
            label = "Marketplace"

    return os.path.join(OUTPUT_DIR, f"{sheet_name}_Validation_{label}.xlsx")


def validate_status_counts(df: pd.DataFrame, sheet_name: str) -> None:
    """
    Parse the 'recommendation' column, collect every status value,
    then validate that DENY and OKAY counts are within [STATUS_MIN, STATUS_MAX].
    """
    if RECOMMENDATION_COL not in df.columns:
        print(f"  [{sheet_name}] ⚠️  '{RECOMMENDATION_COL}' column not found — skipping validation.")
        return

    status_counts: dict[str, int] = {}

    for cell_value in df[RECOMMENDATION_COL]:
        parsed = parse_recommendation(cell_value)
        if not parsed:
            continue
        recommendations = parsed.get("recommendation", [])
        if not isinstance(recommendations, list):
            continue
        for rec in recommendations:
            status = str(rec.get("status", "UNKNOWN")).upper()
            status_counts[status] = status_counts.get(status, 0) + 1

    print(f"\n  [{sheet_name}] 📊 Status counts (from '{RECOMMENDATION_COL}' column):")

    for status in ("DENY", "OKAY"):
        count = status_counts.get(status, 0)
        within_max = count <= STATUS_MAX
        flag = "✅" if within_max else "❌"
        print(
            f"    {flag}  {status}: {count}  "
            f"(max allowed: {STATUS_MAX}, "
            f"{'VALID' if within_max else 'EXCEEDS LIMIT'})"
        )



def apply_yellow_fill(filepath: str) -> None:
    """Paint every cell in every sheet of the given workbook yellow."""
    wb = load_workbook(filepath)
    yellow_fill = PatternFill(
        start_color=YELLOW_HEX, end_color=YELLOW_HEX, fill_type="solid"
    )
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.fill = yellow_fill
    wb.save(filepath)


def process_workbook(input_file: str) -> None:
    """
    Main processing function.
    Each target sheet is saved as its own output file:
        {SHEET_NAME}_Validation_Marketplace.xlsx  or
        {SHEET_NAME}_Validation_Medicaid.xlsx
    """
    # 1. Load all sheets at once
    try:
        all_sheets: dict[str, pd.DataFrame] = pd.read_excel(
            input_file, sheet_name=None, dtype=str
        )
    except FileNotFoundError:
        print(f"❌  Error: '{input_file}' not found.")
        return

    sheets_to_process = list(all_sheets.keys()) if not TARGET_SHEETS else TARGET_SHEETS

    # Warn about sheets that don't exist in the file
    missing = [s for s in sheets_to_process if s not in all_sheets]
    if missing:
        print(f"⚠️   These sheets were not found and will be skipped: {missing}")
    sheets_to_process = [s for s in sheets_to_process if s in all_sheets]

    print(f"\n📂  Input : {input_file}")
    print(f"📁  Output: {os.path.abspath(OUTPUT_DIR)}\n")

    # 2. Process each sheet individually
    for sheet_name in sheets_to_process:
        df = all_sheets[sheet_name].copy()

        # ── a. Clear / set the api response column ──────────────────────
        if API_RESPONSE_COL in df.columns:
            df[API_RESPONSE_COL] = MATCH_VALUE
        else:
            print(f"  [{sheet_name}] ℹ️   '{API_RESPONSE_COL}' column not found — skipping column update.")

        # ── b. Determine output filename ─────────────────────────────────
        output_file = get_output_filename(sheet_name, df)

        # ── c. Save to its own Excel file ────────────────────────────────
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        # ── d. Apply yellow fill via openpyxl ────────────────────────────
        apply_yellow_fill(output_file)

        # ── e. Validate recommendation status counts ─────────────────────
        validate_status_counts(df, sheet_name)

        print(f"  [{sheet_name}] ✅  Saved → '{output_file}'\n")

    print("🎉  All done!")


# ─────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────
if __name__ == "__main__":
    process_workbook(INPUT_FILE)
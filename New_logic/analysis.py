import os
import time
import random
import pandas as pd
import numpy as np
from google import genai
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from dotenv import load_dotenv

# ---------------------------
# CONFIGURATION
# ---------------------------
load_dotenv()
API_KEY = os.getenv("GOOGLE_API_KEY")

if not API_KEY:
    API_KEY = ''
    # raise EnvironmentError("GOOGLE_API_KEY not set. Please add it to your .env file.")

# Initialise the new google-genai client
_genai_client = genai.Client(api_key=API_KEY)
EMBED_MODEL = "models/gemini-embedding-001"

# ---------------------------
# INPUT CONFIGURATION
#
# All entries in src_obj  → merged into one DataFrame d1
# All entries in target_obj → merged into one DataFrame d2
#
# Each entry describes one quarter's Excel file:
#   - year        : int  – the year of the data
#   - quarter     : str  – e.g. "Q3", "Q4"
#   - file_name   : str  – path to the .xlsx file
#   - sheet_names : list – sheet names inside that Excel file to load
#
# Comparison model (first-match):
#   For every row in d2, scan d1 for the first row with a matching Code.
#   When a match is found → compare and break (move to next d2 row).
# ---------------------------
src_obj = [
    {
        "year": 2025,
        "quarter": "Q4",
        "file_name": "/Users/xaiuser/Local_Projects/Code_analysis/New_logic/auth_matrix_files/Q4_2025.xlsx",
        "sheet_names": ["Medicaid","WA"],
    },
    {
        "year": 2025,
        "quarter": "Q3",
        "file_name": "/Users/xaiuser/Local_Projects/Code_analysis/New_logic/auth_matrix_files/Q3_2025.xlsx",
        "sheet_names": ["Medicaid", "WA"],
    },
        {
        "year": 2026,
        "quarter": "Q1",
        "file_name": "/Users/xaiuser/Local_Projects/Code_analysis/New_logic/auth_matrix_files/Q1_2026.xlsx",
        "sheet_names": ["Medicaid", "WA"],
    }
]

target_obj = [
    {
        "year": 2025,
        "quarter": "Q2",
        "file_name": "/Users/xaiuser/Local_Projects/Code_analysis/New_logic/auth_matrix_files/Q2_2025.xlsx",
        "sheet_names": ["Medicaid","WA"],
    }
]

# Column label metadata
SRC_LABEL = "d1 (src)"
TGT_LABEL = "d2 (tgt)"

# ---------------------------
# GLOBAL EMBEDDING CACHE
# ---------------------------
_embedding_cache: dict = {}

# ---------------------------
# CANDIDATE COLUMN NAMES
# ---------------------------
CODE_CANDIDATES = ["Code", "CPT Code", "Procedure Code", "Service Code", "HCPCS"]
TEXT_CANDIDATES = [
    "Code Notes", "MHI Code Notes", "Notes",
    "Code Description", "Service Description",
]
MEDICAID_CANDIDATES = ["Medicaid", "MHI Medicaid", "Medicaid PA"]


# ---------------------------
# HELPER UTILITIES
# ---------------------------
def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Strip whitespace / newlines from column names."""
    df.columns = df.columns.str.strip().str.replace("\n", " ").str.replace("\r", " ")
    return df


def find_column(df: pd.DataFrame, candidates: list):
    """Return the first df column whose lowercased name is in *candidates*."""
    lower = [c.lower() for c in candidates]
    for col in df.columns:
        if col.lower().strip() in lower:
            return col
    return None


def retry_with_backoff(retries: int = 5, backoff_in_seconds: int = 2):
    """Decorator – exponential-backoff retry for flaky API calls."""
    def decorator(func):
        def wrapper(*args, **kwargs):
            attempt = 0
            while True:
                try:
                    return func(*args, **kwargs)
                except Exception as exc:
                    if attempt == retries:
                        print(f"    [WARN] API call failed after {retries} retries: {exc}")
                        raise
                    sleep_time = (backoff_in_seconds * 2 ** attempt) + random.uniform(0, 1)
                    time.sleep(sleep_time)
                    attempt += 1
        return wrapper
    return decorator


@retry_with_backoff(retries=5, backoff_in_seconds=2)
def _api_embed(text: str) -> np.ndarray:
    time.sleep(0.4)
    resp = _genai_client.models.embed_content(
        model=EMBED_MODEL,
        contents=text,
        config={"task_type": "semantic_similarity"},
    )
    return np.array(resp.embeddings[0].values)


def embed(text) -> np.ndarray:
    """Cached embedding: API is only called on a cache miss."""
    key = str(text).strip() if text else "empty"
    if not key:
        key = "empty"
    if key not in _embedding_cache:
        try:
            _embedding_cache[key] = _api_embed(key)
        except Exception:
            _embedding_cache[key] = np.zeros(768)
    return _embedding_cache[key]


def cosine_similarity(v1: np.ndarray, v2: np.ndarray) -> float:
    n1, n2 = np.linalg.norm(v1), np.linalg.norm(v2)
    if n1 == 0 or n2 == 0:
        return 0.0
    return float(np.dot(v1, v2) / (n1 * n2))


def clean(val) -> str:
    """Normalise a cell value to a plain string; treat nan/None as ''."""
    s = str(val).strip()
    return "" if s in ("nan", "None", "NaN") else s


# ---------------------------
# DATA LOADING
# ---------------------------
def load_and_merge(entries: list[dict]) -> pd.DataFrame:
    """
    Load every requested sheet from every entry in *entries* and
    concatenate them all into a single DataFrame.

    Adds helper columns:
      __year    – year from config
      __quarter – quarter from config
      __sheet   – sheet tab name
      Code      – normalised code value
      __text    – normalised text/notes value
    """
    frames = []

    for entry in entries:
        file_path = entry["file_name"]
        year      = entry["year"]
        quarter   = entry["quarter"]
        sheets    = entry["sheet_names"]

        if not os.path.isfile(file_path):
            print(f"  [ERROR] File not found: {file_path}  (entry {year} {quarter})")
            continue

        xls = pd.ExcelFile(file_path)

        for sheet in sheets:
            if sheet not in xls.sheet_names:
                print(f"  [WARN]  Sheet '{sheet}' not in {file_path} "
                      f"(available: {xls.sheet_names})")
                continue

            df = normalize_columns(pd.read_excel(xls, sheet_name=sheet))
            df = df.replace(np.nan, "", regex=True)

            # Detect Code column
            code_col = find_column(df, CODE_CANDIDATES)
            if not code_col:
                print(f"  [WARN]  No Code column in sheet '{sheet}' of {file_path} – skipping.")
                continue

            df["Code"] = df[code_col].astype(str).str.strip()

            # Detect notes/text column (optional)
            text_col = find_column(df, TEXT_CANDIDATES)
            df["__text"] = df[text_col].apply(clean) if text_col else ""

            # Detect Medicaid column (optional)
            med_col = find_column(df, MEDICAID_CANDIDATES)
            df["__medicaid"] = df[med_col].apply(clean) if med_col else ""

            df["__year"]    = year
            df["__quarter"] = quarter
            df["__sheet"]   = sheet

            frames.append(df)

    if not frames:
        return pd.DataFrame()

    merged = pd.concat(frames, ignore_index=True)
    return merged


# ---------------------------
# CORE COMPARISON (first-match)
# ---------------------------
def compare_d2_against_d1(d1: pd.DataFrame, d2: pd.DataFrame) -> pd.DataFrame:
    """
    For every row in d2:
      - Search d1 for the first row with the same Code.
      - If found  → compare text; record result; move to next d2 row.
      - If not found → mark as 'New in target'.

    Codes in d1 that were never matched by any d2 row → 'Removed from target'.

    Returns a report DataFrame.
    """
    if d1.empty or d2.empty:
        print("  [WARN] One of the merged DataFrames is empty – nothing to compare.")
        return pd.DataFrame()

    # Build a dict: code → first matching d1 row  (first-match semantics)
    d1_lookup: dict[str, dict] = {}
    for _, row in d1.iterrows():
        code = row["Code"]
        if code not in d1_lookup:               # keep only the FIRST occurrence
            d1_lookup[code] = row.to_dict()

    matched_d1_codes: set[str] = set()          # track which d1 codes were hit
    report_rows = []

    # ---- Iterate d2 rows ----
    for _, tgt_row in d2.iterrows():
        code     = tgt_row["Code"]
        tgt_text = clean(tgt_row.get("__text", ""))
        tgt_med  = clean(tgt_row.get("__medicaid", ""))
        tgt_meta = f"{tgt_row.get('__year','')} {tgt_row.get('__quarter','')} / {tgt_row.get('__sheet','')}"

        # Build combined value string for d2
        tgt_val_str = f"Notes: {tgt_text}"
        if tgt_med:
            tgt_val_str += f" | Medicaid: {tgt_med}"

        if code not in d1_lookup:
            report_rows.append({
                "Code":       code,
                "Status":     "New in Target",
                "Severity":   "New Entry",
                "d1 Source":  "",
                "d1 Value":   "",
                "d2 Source":  tgt_meta,
                "d2 Value":   tgt_val_str,
                "Similarity": "",
            })
            continue

        # --- First match found in d1 ---
        src_row  = d1_lookup[code]
        src_text = clean(src_row.get("__text", ""))
        src_med  = clean(src_row.get("__medicaid", ""))
        src_meta = f"{src_row.get('__year','')} {src_row.get('__quarter','')} / {src_row.get('__sheet','')}"

        # Build combined value string for d1
        src_val_str = f"Notes: {src_text}"
        if src_med:
            src_val_str += f" | Medicaid: {src_med}"

        matched_d1_codes.add(code)

        notes_same = (src_text == tgt_text)
        med_same   = (src_med  == tgt_med)

        if notes_same and med_same:
            report_rows.append({
                "Code":       code,
                "Status":     "No Change",
                "Severity":   "No Change",
                "d1 Source":  src_meta,
                "d1 Value":   src_val_str,
                "d2 Source":  tgt_meta,
                "d2 Value":   tgt_val_str,
                "Similarity": 1.0,
            })
        else:
            # --- Build combined severity (mirrors compare_analyse.py) ---
            severity_parts = []

            # 1. Medicaid change (string equality check)
            if not med_same:
                severity_parts.append("Medicaid Change")

            # 2. Notes change (semantic similarity)
            sim = 0.0
            if not notes_same:
                if src_text and tgt_text:
                    sim = cosine_similarity(embed(src_text), embed(tgt_text))
                if not src_text or not tgt_text:
                    severity_parts.append("Severe Change")
                elif sim < 0.55:
                    severity_parts.append("Severe Change")
                elif sim < 0.80:
                    severity_parts.append("Moderate Change")
                else:
                    severity_parts.append("Minor Wording Change")

            report_rows.append({
                "Code":       code,
                "Status":     "Modified",
                "Severity":   "; ".join(severity_parts),
                "d1 Source":  src_meta,
                "d1 Value":   src_val_str,
                "d2 Source":  tgt_meta,
                "d2 Value":   tgt_val_str,
                "Similarity": round(sim, 4) if sim else "",
            })

    # ---- Codes in d1 never matched by any d2 row → Removed ----
    removed_codes = set(d1_lookup.keys()) - matched_d1_codes
    for code in removed_codes:
        src_row     = d1_lookup[code]
        src_text    = clean(src_row.get("__text", ""))
        src_med     = clean(src_row.get("__medicaid", ""))
        src_meta    = f"{src_row.get('__year','')} {src_row.get('__quarter','')} / {src_row.get('__sheet','')}"
        src_val_str = f"Notes: {src_text}"
        if src_med:
            src_val_str += f" | Medicaid: {src_med}"
        report_rows.append({
            "Code":       code,
            "Status":     "Removed from Target",
            "Severity":   "Severe Change",
            "d1 Source":  src_meta,
            "d1 Value":   src_val_str,
            "d2 Source":  "",
            "d2 Value":   "",
            "Similarity": "",
        })

    return pd.DataFrame(
        report_rows,
        columns=["Code", "Status", "Severity",
                 "d1 Source", "d1 Value",
                 "d2 Source", "d2 Value",
                 "Similarity"],
    )


# ---------------------------
# EXCEL OUTPUT HELPERS
# ---------------------------
def compute_summary(df: pd.DataFrame) -> dict:
    return {
        "Total d1 (src) codes":    len(df[df["d1 Value"] != ""]),
        "Total d2 (tgt) codes":    len(df[df["d2 Value"] != ""]),
        "No Change":               (df["Status"] == "No Change").sum(),
        "Modified":                (df["Status"] == "Modified").sum(),
        "Medicaid Change":         df["Severity"].str.contains("Medicaid Change",       na=False).sum(),
        "Severe Change":           df["Severity"].str.contains("Severe Change",         na=False).sum(),
        "Moderate Change":         df["Severity"].str.contains("Moderate Change",       na=False).sum(),
        "Minor Wording Change":    df["Severity"].str.contains("Minor Wording Change",  na=False).sum(),
        "New in Target":           (df["Status"] == "New in Target").sum(),
        "Removed from Target":     (df["Status"] == "Removed from Target").sum(),
    }


def write_summary(ws, summary: dict) -> None:
    # +2 = 1 for the "SUMMARY" title row + 1 blank gap before the data table
    insert_count = len(summary) + 2
    ws.insert_rows(1, amount=insert_count)
    ws["A1"] = "SUMMARY"
    ws["A1"].font = Font(bold=True)
    for i, (key, val) in enumerate(summary.items(), start=2):
        ws[f"A{i}"] = key
        ws[f"B{i}"] = val


def apply_conditional_formatting(ws) -> None:
    """
    Colour every data row based on Status (col B) with Severity (col C)
    used for sub-shading on Modified rows.

    Colour legend:
      No Change                      → light gray   #F2F2F2
      Modified / Medicaid Change     → purple       #E1D3F6
      Modified / Severe Change       → red          #FFC7CE
      Modified / Moderate Change     → yellow       #FFEB9C
      Modified / Minor Wording       → green        #C6EFCE
      Modified / Medicaid+Severe     → red          #FFC7CE  (notes take priority)
      New in Target                  → blue         #BDD7EE
      Removed from Target            → orange       #FFD7A0
    """
    STATUS_COLOURS = {
        "No Change":           "F2F2F2",  # light gray
        "New in Target":       "BDD7EE",  # blue
        "Removed from Target": "FFD7A0",  # orange
    }
    # Order matters: when severity is a combined string like
    # "Medicaid Change; Severe Change", the FIRST match wins.
    SEVERITY_PRIORITY = [
        ("Severe Change",        "FFC7CE"),  # red      — highest priority
        ("Moderate Change",      "FFEB9C"),  # yellow
        ("Minor Wording Change", "C6EFCE"),  # green
        ("Medicaid Change",      "E1D3F6"),  # purple   — lowest priority (alone)
    ]

    header_row   = None
    status_col   = 2
    severity_col = 3

    # Scan for the actual data header row (row containing "Code" in col A)
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Code":
            header_row = r
            break

    if header_row is None:
        return  # nothing to format

    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if val == "Status":
            status_col = col_idx
        elif val == "Severity":
            severity_col = col_idx

    for row in range(header_row + 1, ws.max_row + 1):
        status_val   = str(ws.cell(row=row, column=status_col).value   or "")
        severity_val = str(ws.cell(row=row, column=severity_col).value or "")
        fill_color   = None

        if status_val in STATUS_COLOURS:
            fill_color = STATUS_COLOURS[status_val]
        elif status_val == "Modified":
            for label, hex_c in SEVERITY_PRIORITY:
                if label in severity_val:
                    fill_color = hex_c
                    break

        if fill_color:
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col_idx).fill = fill




# ---------------------------
# MAIN ENTRY POINT
# ---------------------------
def run_analysis() -> None:
    output_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "comparison_output.xlsx",
    )

    print("=" * 60)
    print("  d1 vs d2 Analysis  (first-match on Code)")
    print("=" * 60)

    # --- Build d1: merge all src entries ---
    print("\n[1/3] Loading d1 (src_obj) …")
    d1 = load_and_merge(src_obj)
    print(f"      d1 total rows : {len(d1)}")
    print(f"      d1 unique codes: {d1['Code'].nunique() if not d1.empty else 0}")

    # --- Build d2: merge all target entries ---
    print("\n[2/3] Loading d2 (target_obj) …")
    d2 = load_and_merge(target_obj)
    print(f"      d2 total rows : {len(d2)}")
    print(f"      d2 unique codes: {d2['Code'].nunique() if not d2.empty else 0}")

    # --- Compare ---
    print("\n[3/3] Comparing d2 rows against d1 (first-match per Code) …")
    report = compare_d2_against_d1(d1, d2)

    if report.empty:
        print("\n[INFO] No comparison results. Check file paths and sheet names.")
        return

    report = report.replace(np.nan, "", regex=True)
    summary = compute_summary(report)

    # --- Write Excel ---
    print(f"\nWriting output → {output_path}")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        report.to_excel(writer, sheet_name="d1 vs d2 Comparison", index=False)

    wb = load_workbook(output_path)
    ws = wb["d1 vs d2 Comparison"]
    write_summary(ws, summary)
    apply_conditional_formatting(ws)
    wb.save(output_path)

    # --- Console summary ---
    print("\n🎉 Analysis complete!")
    print(f"   Output  : {output_path}")
    print(f"   Sheet   : 'd1 vs d2 Comparison'  ({len(report)} rows)")
    print("\nSummary:")
    for k, v in summary.items():
        print(f"  {k:<30} {v}")


if __name__ == "__main__":
    run_analysis()

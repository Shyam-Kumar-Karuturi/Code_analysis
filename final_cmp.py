import os
import time
import random
import pandas as pd
import numpy as np
import google.generativeai as genai
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from dotenv import load_dotenv

# ---------------------------
# CONFIGURATION
# ---------------------------
load_dotenv()
API_KEY = os.getenv("GOOGLE_API_KEY") or 'AIzaSyD9WvnD_GR_rJcU2iZcjXlhtILRlbx5n_c'

if not API_KEY:
    print("CRITICAL WARNING: No Google API Key found. AI features will fail.")

genai.configure(api_key=API_KEY)
EMBED_MODEL = "models/text-embedding-004"

INPUT_FILE = "Merged_25Q4_26Q1_Analysis_3.xlsx"
OUTPUT_FILE = "Final_Comparison_Report.xlsx"

# Global Cache to speed up AI processing
_embedding_cache = {}


# ---------------------------
# HELPERS
# ---------------------------
def clean_val(val):
    """Converts input to string, removing 'nan', 'none', and whitespace."""
    if val is None: return ""
    s = str(val).strip()
    if s.lower() in ['nan', 'none', '']: return ""
    return s


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Strip whitespace/newlines from column names."""
    df.columns = df.columns.str.strip().str.replace('\n', ' ').str.replace('\r', ' ')
    return df


def find_column(df: pd.DataFrame, possible_names):
    """Find column name in df that matches one of possible_names (case-insensitive)."""
    lower_targets = [p.lower() for p in possible_names]
    for col in df.columns:
        if col.lower().strip() in lower_targets:
            return col
    return None


def retry_with_backoff(retries=3, backoff_in_seconds=1):
    def decorator(func):
        def wrapper(*args, **kwargs):
            x = 0
            while True:
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    if x == retries:
                        return np.zeros(768)
                    time.sleep((backoff_in_seconds * 2 ** x) + random.uniform(0, 1))
                    x += 1

        return wrapper

    return decorator


@retry_with_backoff(retries=3, backoff_in_seconds=1)
def generate_embedding_from_api(text: str):
    time.sleep(0.4)
    resp = genai.embed_content(model=EMBED_MODEL, content=str(text), task_type="semantic_similarity")
    return np.array(resp["embedding"])


def embed(text: str):
    if not text: return np.zeros(768)
    text_str = clean_val(text)
    if text_str == "": return np.zeros(768)

    if text_str in _embedding_cache:
        return _embedding_cache[text_str]

    try:
        vector = generate_embedding_from_api(text_str)
        _embedding_cache[text_str] = vector
        return vector
    except Exception:
        return np.zeros(768)


def cosine_similarity(v1, v2):
    n1 = np.linalg.norm(v1)
    n2 = np.linalg.norm(v2)
    if n1 == 0 or n2 == 0: return 0.0
    return float(np.dot(v1, v2) / (n1 * n2))


# ---------------------------
# CORE LOGIC
# ---------------------------
def compare_notes_and_medicaid(q3_df: pd.DataFrame, q4_df: pd.DataFrame, notes_col: str, medicaid_col: str = None):
    q3_df = q3_df.copy()
    q4_df = q4_df.copy()

    # Normalize 'Code' column
    code_col_q3 = find_column(q3_df, ["Code", "CPT Code", "Procedure Code", "Service Code", "HCPCS"])
    code_col_q4 = find_column(q4_df, ["Code", "CPT Code", "Procedure Code", "Service Code", "HCPCS"])

    if not code_col_q3 or not code_col_q4: return None

    # Standardize Codes
    q3_df["Code"] = q3_df[code_col_q3].astype(str).str.strip()
    q4_df["Code"] = q4_df[code_col_q4].astype(str).str.strip()

    # --- CRITICAL FIX: DEDUPLICATE CODES ---
    # Keeps the first occurrence if duplicates exist to prevent "Index must be unique" error
    q3_df = q3_df.drop_duplicates(subset=["Code"], keep="first")
    q4_df = q4_df.drop_duplicates(subset=["Code"], keep="first")

    # Create lookup dictionary
    q4_lookup = q4_df.set_index("Code").to_dict('index')

    report_rows = []
    has_med_col = medicaid_col is not None

    # 1. Process OLD codes
    for _, row in q3_df.iterrows():
        code = row["Code"]
        q3_notes = clean_val(row.get(notes_col)) if notes_col else ""
        q3_med = clean_val(row.get(medicaid_col)) if has_med_col else ""

        q3_val_str = f"Notes: {q3_notes}"
        if has_med_col: q3_val_str += f" | Medicaid: {q3_med}"

        if code not in q4_lookup:
            report_rows.append({
                "Code": code, "Status": "Removed in 2026 Q1",
                "Severity": "Severe Change",
                "2025 Q4 Value": q3_val_str, "2026 Q1 Value": ""
            })
            continue

        row_new = q4_lookup[code]
        q4_notes = clean_val(row_new.get(notes_col)) if notes_col else ""
        q4_med = clean_val(row_new.get(medicaid_col)) if has_med_col else ""

        q4_val_str = f"Notes: {q4_notes}"
        if has_med_col: q4_val_str += f" | Medicaid: {q4_med}"

        notes_same = (q3_notes == q4_notes)
        med_same = (q3_med == q4_med) if has_med_col else True

        if notes_same and med_same:
            report_rows.append({
                "Code": code, "Status": "No Change",
                "Severity": "No Change",
                "2025 Q4 Value": q3_val_str, "2026 Q1 Value": q4_val_str
            })
        else:
            severity_parts = []
            if has_med_col and not med_same: severity_parts.append("Medicaid Change")

            if not notes_same:
                sim = cosine_similarity(embed(q3_notes), embed(q4_notes))
                if sim < 0.6:
                    severity_parts.append("Severe Change")
                elif sim < 0.85:
                    severity_parts.append("Moderate Change")
                else:
                    severity_parts.append("Minor Wording Change")

            report_rows.append({
                "Code": code, "Status": "Modified",
                "Severity": "; ".join(severity_parts),
                "2025 Q4 Value": q3_val_str, "2026 Q1 Value": q4_val_str
            })

    # 2. Process NEW codes
    q3_codes = set(q3_df["Code"])
    new_codes = set(q4_lookup.keys()) - q3_codes

    for code in new_codes:
        row_new = q4_lookup[code]
        q4_notes = clean_val(row_new.get(notes_col)) if notes_col else ""
        q4_med = clean_val(row_new.get(medicaid_col)) if has_med_col else ""

        q4_val_str = f"Notes: {q4_notes}"
        if has_med_col: q4_val_str += f" | Medicaid: {q4_med}"

        report_rows.append({
            "Code": code, "Status": "New in 2026 Q1",
            "Severity": "New Entry",
            "2025 Q4 Value": "", "2026 Q1 Value": q4_val_str
        })

    return pd.DataFrame(report_rows, columns=["Code", "Status", "Severity", "2025 Q4 Value", "2026 Q1 Value"])


# ---------------------------
# MAIN EXECUTION
# ---------------------------
def main():
    print("=" * 60)
    print("       🚀 RUNNING FINAL ANALYSIS (FIXED)")
    print("=" * 60)

    if not os.path.exists(INPUT_FILE):
        print(f"❌ Error: Input file '{INPUT_FILE}' not found.")
        return

    # 1. READ INPUT
    print(f"📂 Reading: {INPUT_FILE}...")
    try:
        try:
            import python_calamine
            engine = 'calamine'
            print("   (Using fast 'calamine' engine)")
        except ImportError:
            engine = 'openpyxl'
            print("   (Using standard engine)")

        xls = pd.ExcelFile(INPUT_FILE, engine=engine)
    except Exception as e:
        print(f"❌ Error reading Excel: {e}")
        return

    # 2. DETECT PAIRS
    sheet_names = xls.sheet_names
    states = set()
    for s in sheet_names:
        if "25Q4" in s or "26Q1" in s:
            parts = s.split(" ")
            states.add(parts[0])

    print(f"✅ Found {len(states)} potential states: {sorted(list(states))}")

    # 3. PROCESS & WRITE
    print(f"\n💾 Analyzing and writing to: {OUTPUT_FILE}...")

    with pd.ExcelWriter(OUTPUT_FILE, engine='xlsxwriter') as writer:
        for state in sorted(list(states)):
            sheet_old = f"{state} 25Q4"
            sheet_new = f"{state} 26Q1"
            sheet_out = f"{state} 25Q4 vs 26Q1"

            if sheet_old not in sheet_names or sheet_new not in sheet_names: continue

            print(f"   Processing: {state}...", end=" ", flush=True)

            try:
                df_old = pd.read_excel(xls, sheet_name=sheet_old, dtype=str)
                df_new = pd.read_excel(xls, sheet_name=sheet_new, dtype=str)

                df_old = normalize_columns(df_old)
                df_new = normalize_columns(df_new)

                notes_candidates = ["Code Notes", "Notes", "Additional Notes", "Comments"]
                med_candidates = ["Medicaid", "MHI Medicaid", "Medicaid PA"]

                notes_col = find_column(df_old, notes_candidates)
                med_col = find_column(df_old, med_candidates)

                df_result = compare_notes_and_medicaid(df_old, df_new, notes_col, med_col)

                if df_result is not None:
                    df_result.to_excel(writer, sheet_name=sheet_out, index=False)
                    print("Done ✅")
                else:
                    print("Skipped (Missing 'Code' column) ⚠️")

            except Exception as e:
                print(f"Error: {e} ❌")

    # 4. FORMATTING & SUMMARY
    print("\n🎨 Applying formatting and generating summaries...")
    wb = load_workbook(OUTPUT_FILE)

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

    for ws in wb.worksheets:
        if "vs" in ws.title:
            # Counters
            total = 0
            changes = 0
            new_codes = 0
            removed = 0

            # Identify columns
            rows = list(ws.rows)
            if not rows: continue

            header = [c.value for c in rows[0]]
            try:
                status_idx = header.index("Status")
                sev_idx = header.index("Severity")

                # Iterate rows (skip header)
                for row in rows[1:]:
                    total += 1
                    status_val = str(row[status_idx].value)
                    sev_val = str(row[sev_idx].value)

                    # Colors
                    if "Severe" in sev_val or "Removed" in status_val:
                        for c in row: c.fill = red_fill
                        if "Removed" in status_val:
                            removed += 1
                        else:
                            changes += 1
                    elif "Moderate" in sev_val:
                        for c in row: c.fill = yellow_fill
                        changes += 1
                    elif "Minor" in sev_val:
                        for c in row: c.fill = green_fill
                        changes += 1
                    elif "New" in status_val:
                        for c in row: c.fill = blue_fill
                        new_codes += 1
                    elif "Modified" in status_val:
                        changes += 1
            except ValueError:
                pass

            # --- WRITE SUMMARY AT TOP ---
            ws.insert_rows(1, amount=6)

            # Title
            ws["A1"] = "COMPARISON SUMMARY"
            ws["A1"].font = Font(bold=True, size=14)

            # Stats
            ws["A2"] = "Total Codes Processed:"
            ws["B2"] = total

            ws["A3"] = "Total Changes (Modified):"
            ws["B3"] = changes

            ws["A4"] = "New Codes (Added):"
            ws["B4"] = new_codes

            ws["A5"] = "Removed Codes:"
            ws["B5"] = removed

            # Styling
            for row in range(2, 6):
                ws[f"A{row}"].font = Font(bold=True)
                ws[f"B{row}"].alignment = Alignment(horizontal='left')

    wb.save(OUTPUT_FILE)
    print(f"✅ Success! Report saved to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
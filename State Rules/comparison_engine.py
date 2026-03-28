import os
import re
import time
import random
import pandas as pd
import numpy as np
from google import genai
from google.genai import types as genai_types
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from dotenv import load_dotenv

# ---------------------------
# CONFIGURATION
# ---------------------------
load_dotenv()
API_KEY = os.getenv("GOOGLE_API_KEY")

if not API_KEY:
    API_KEY = 'AIzaSyBLQ0ATKyuJPU_r7b7PBu5C0JWNAJJNlPA'

client = genai.Client(api_key=API_KEY)
EMBED_MODEL = "models/gemini-embedding-001"

_embedding_cache = {}

def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = df.columns.str.strip().str.replace('\n', ' ').str.replace('\r', ' ')
    return df

_INVISIBLE = re.compile(r'[\xa0\u200b\u200c\u200d\ufeff\u00ad\u2060]')

def normalize_code(val) -> str:
    """
    Aggressively normalise a code value so that invisible characters,
    non-breaking spaces, BOM markers, and leading/trailing whitespace
    never prevent two visually-identical codes from matching.
    Also upper-cases the result for case-insensitive matching.
    """
    if val is None:
        return ""
    s = str(val)
    s = _INVISIBLE.sub('', s)        # strip invisible unicode chars
    s = re.sub(r'\s+', ' ', s).strip()  # collapse whitespace
    return s.upper()

def find_column(df: pd.DataFrame, possible_names):
    lower_targets = [p.lower() for p in possible_names]
    for col in df.columns:
        if col.lower().strip() in lower_targets:
            return col
    return None

def retry_with_backoff(retries=5, backoff_in_seconds=2):
    def decorator(func):
        def wrapper(*args, **kwargs):
            x = 0
            while True:
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    if x == retries:
                        print(f"    Failed after {retries} retries: {e}")
                        raise e
                    sleep = (backoff_in_seconds * 2 ** x) + random.uniform(0, 1)
                    time.sleep(sleep)
                    x += 1
        return wrapper
    return decorator

@retry_with_backoff(retries=5, backoff_in_seconds=2)
def generate_embedding_from_api(text: str):
    time.sleep(0.4)
    resp = client.models.embed_content(
        model=EMBED_MODEL,
        contents=str(text),
        config=genai_types.EmbedContentConfig(task_type="SEMANTIC_SIMILARITY")
    )
    return np.array(resp.embeddings[0].values)

def embed(text: str):
    if text is None or str(text).strip() == "":
        text = "empty"
    text_str = str(text).strip()
    if text_str in _embedding_cache:
        return _embedding_cache[text_str]
    try:
        vector = generate_embedding_from_api(text_str)
        _embedding_cache[text_str] = vector
        return vector
    except Exception:
        return np.zeros(768)

def cosine_similarity(v1, v2):
    n1 = np.linalg.norm(v1)
    n2 = np.linalg.norm(v2)
    if n1 == 0 or n2 == 0:
        return 0.0
    return float(np.dot(v1, v2) / (n1 * n2))

def _parse_alias(alias: str):
    """Extract (quarter_id, year) from an alias like 'Q3 2025'. Falls back gracefully."""
    import re
    m = re.search(r'(Q[1-4])', alias, re.IGNORECASE)
    y = re.search(r'(20\d{2})', alias)
    quarter_id = m.group(1).upper() if m else ""
    year       = y.group(1)         if y else ""
    return quarter_id, year

def write_summary(ws, summary: dict, source_aliases=None, tgt_alias=None):
    # How many header rows to insert: base 6 + 1 blank before data + extras for aliases
    alias_rows = 0
    if source_aliases: alias_rows += len(source_aliases)
    if tgt_alias:      alias_rows += 1
    insert_count = 7 + alias_rows   # 6 summary + 1 blank
    ws.insert_rows(1, amount=insert_count)

    ws["A1"] = "COMPARISON SUMMARY"
    ws["A1"].font = Font(bold=True, size=14)

    row = 2
    # File alias metadata
    if tgt_alias:
        ws[f"A{row}"] = "Target File"
        ws[f"B{row}"] = tgt_alias
        ws[f"A{row}"].font = Font(bold=True)
        row += 1
    if source_aliases:
        for i, sa in enumerate(source_aliases, start=1):
            ws[f"A{row}"] = f"Source File {i}"
            ws[f"B{row}"] = sa
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

    # Stats
    for key, val in summary.items():
        ws[f"A{row}"] = key
        ws[f"B{row}"] = val
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].alignment = Alignment(horizontal='left')
        row += 1

    # Blank separator row before the data table
    # (row is now pointing at the blank row; next row = data header)
    row += 1   # leave one blank row

def apply_conditional_formatting(ws):
    # Status-based colors (primary)
    status_colors = {
        "Removed in Target": "FF9999",   # red
        "New in Target":     "BDD7EE",   # blue
        "No Change":         "E2EFDA",   # light green
    }
    # Severity sub-colors (used when Status == "Modified")
    severity_colors = {
        "Severe Change":        "FFC7CE",  # red-pink
        "Moderate Change":      "FFEB9C",  # yellow
        "Minor Wording Change": "C6EFCE",  # light green
        "Medicaid Change":      "E1D3F6",  # purple
    }
    default_modified_color = "FFD9B3"     # orange fallback

    # ── Dynamically find the header row ────────────────────────────────────
    header_row = None
    for r in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=r, column=1).value
        if cell_val is not None and str(cell_val).strip().lower() == "code":
            header_row = r
            break
    if header_row is None:
        return   # no header found, skip formatting

    # ── Locate Status and Severity column indices ──────────────────────────
    headers = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]
    status_col_idx   = next((i+1 for i, h in enumerate(headers) if h and "Status"   in str(h)), None)
    severity_col_idx = next((i+1 for i, h in enumerate(headers) if h and "Severity" in str(h)), None)

    # ── Yellow bold header row ─────────────────────────────────────────────
    header_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
    header_font = Font(bold=True)
    for col in range(1, ws.max_column + 1):
        c = ws.cell(row=header_row, column=col)
        c.fill = header_fill
        c.font = header_font

    # ── Color each data row ────────────────────────────────────────────────
    for row in range(header_row + 1, ws.max_row + 1):
        status_val   = str(ws.cell(row=row, column=status_col_idx).value   or "") if status_col_idx   else ""
        severity_val = str(ws.cell(row=row, column=severity_col_idx).value or "") if severity_col_idx else ""

        # Pick fill color
        if status_val in status_colors:
            fill_color = status_colors[status_val]
        elif status_val == "Modified":
            fill_color = default_modified_color
            for sev_key, sev_col in severity_colors.items():
                if sev_key in severity_val:
                    fill_color = sev_col
                    break
        else:
            fill_color = "FFFFFF"   # unknown status → white

        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = fill


def compare_notes_and_medicaid(q3_df: pd.DataFrame, q4_df: pd.DataFrame,
                               notes_col: str, medicaid_col: str = None,
                               source_aliases: list = None, tgt_alias: str = ""):
    q3_df = q3_df.copy()
    q4_df = q4_df.copy()

    code_col_q3 = find_column(q3_df, ["STD_CODE", "Code", "CPT Code", "Procedure Code", "Service Code", "HCPCS"])
    code_col_q4 = find_column(q4_df, ["STD_CODE", "Code", "CPT Code", "Procedure Code", "Service Code", "HCPCS"])

    if not code_col_q3 or not code_col_q4: return None

    # Use normalize_code so that invisible chars / NBSP / case never break matching
    q3_df["Code"] = q3_df[code_col_q3].apply(normalize_code)
    q4_df["Code"] = q4_df[code_col_q4].apply(normalize_code)

    # Drop blank codes (empty string after normalisation)
    q3_df = q3_df[q3_df["Code"] != ""].copy()
    q4_df = q4_df[q4_df["Code"] != ""].copy()

    q3_df = q3_df.drop_duplicates(subset=["Code"], keep="last")
    q4_df = q4_df.drop_duplicates(subset=["Code"], keep="last")

    q4_lookup = q4_df.set_index("Code").to_dict('index')

    # Build a per-code → first-source-alias mapping using the _source_alias column
    # (set by run_gui_analysis before concat)
    code_to_alias = {}       # code → alias of first source that has it
    if "_source_alias" in q3_df.columns:
        for alias_val, grp in q3_df.groupby("_source_alias", sort=False):
            for code in grp["Code"].unique():
                if code not in code_to_alias:
                    code_to_alias[code] = alias_val

    def _tag(alias):
        q, y = _parse_alias(alias) if alias else ("", "")
        return q, y

    tgt_q, tgt_y = _tag(tgt_alias)

    report_rows = []
    has_med_col = medicaid_col is not None

    for _, row in q3_df.iterrows():
        code = row["Code"]
        q3_notes = str(row.get(notes_col)).strip() if notes_col else ""
        if q3_notes == "nan" or q3_notes == "None": q3_notes = ""

        q3_med = str(row.get(medicaid_col)).strip() if has_med_col else ""
        if q3_med == "nan" or q3_med == "None": q3_med = ""

        q3_val_str = f"Notes: {q3_notes}"
        if has_med_col: q3_val_str += f" | Medicaid: {q3_med}"

        if code not in q4_lookup:
            src_alias_for_code = code_to_alias.get(code, "")
            report_rows.append({
                "Code": code, "Status": "Removed in Target",
                "Severity": "Severe Change",
                "Source Value": q3_val_str, "Target Value": "",
                "quarter_id": _tag(src_alias_for_code)[0],
                "year":       _tag(src_alias_for_code)[1],
            })
            continue

        row_new = q4_lookup[code]
        q4_notes = str(row_new.get(notes_col)).strip() if notes_col else ""
        if q4_notes == "nan" or q4_notes == "None": q4_notes = ""

        q4_med = str(row_new.get(medicaid_col)).strip() if has_med_col else ""
        if q4_med == "nan" or q4_med == "None": q4_med = ""

        q4_val_str = f"Notes: {q4_notes}"
        if has_med_col: q4_val_str += f" | Medicaid: {q4_med}"

        notes_same = (q3_notes == q4_notes)
        med_same = (q3_med == q4_med) if has_med_col else True

        if notes_same and med_same:
            report_rows.append({
                "Code": code, "Status": "No Change",
                "Severity": "No Change",
                "Source Value": q3_val_str, "Target Value": q4_val_str,
                "quarter_id": _tag(code_to_alias.get(code, tgt_alias))[0],
                "year":       _tag(code_to_alias.get(code, tgt_alias))[1],
            })
        else:
            severity_parts = []
            if has_med_col and not med_same: severity_parts.append("Medicaid Change")

            if not notes_same:
                sim = cosine_similarity(embed(q3_notes), embed(q4_notes))
                if sim < 0.6: severity_parts.append("Severe Change")
                elif sim < 0.85: severity_parts.append("Moderate Change")
                else: severity_parts.append("Minor Wording Change")

            report_rows.append({
                "Code": code, "Status": "Modified",
                "Severity": "; ".join(severity_parts),
                "Source Value": q3_val_str, "Target Value": q4_val_str,
                "quarter_id": _tag(code_to_alias.get(code, tgt_alias))[0],
                "year":       _tag(code_to_alias.get(code, tgt_alias))[1],
            })

    q3_codes = set(q3_df["Code"])
    new_codes = set(q4_lookup.keys()) - q3_codes

    for code in new_codes:
        row_new = q4_lookup[code]
        q4_notes = str(row_new.get(notes_col)).strip() if notes_col else ""
        if q4_notes == "nan" or q4_notes == "None": q4_notes = ""

        q4_med = str(row_new.get(medicaid_col)).strip() if has_med_col else ""
        if q4_med == "nan" or q4_med == "None": q4_med = ""

        q4_val_str = f"Notes: {q4_notes}"
        if has_med_col: q4_val_str += f" | Medicaid: {q4_med}"

        report_rows.append({
            "Code": code, "Status": "New in Target",
            "Severity": "New Entry",
            "Source Value": "", "Target Value": q4_val_str,
            "quarter_id": tgt_q, "year": tgt_y,
        })

    return pd.DataFrame(report_rows, columns=["Code", "Status", "Severity",
                                              "Source Value", "Target Value",
                                              "quarter_id", "year"])

def find_header_row(file_path, sheet_name):
    try:
        df_preview = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=200, dtype=str)
    except Exception:
        return 0

    target_headers = ["code", "cpt code", "procedure code", "hcpcs", "service code"]

    for idx, row in df_preview.iterrows():
        row_values = [str(x).strip().lower() for x in row.values]
        if any(t in row_values for t in target_headers):
            return idx

    return 0

def sanitize_sheet_name(name: str) -> str:
    """Ensure name is a valid Excel sheet name (max 31 chars, no forbidden chars)."""
    forbidden = r'[]:*?/\\'
    for ch in forbidden:
        name = name.replace(ch, '')
    return name[:31].strip()

def run_gui_analysis(mappings, output_file, progress_callback=None):
    """
    mappings: list of dicts:
     [
       {"src_file": path, "src_sheet": name, "tgt_file": path, "tgt_sheet": name, "out_name": name}
     ]
    """
    if progress_callback: progress_callback("Starting Analysis...")
    all_sheets_data = {}

    code_candidates = ["Code", "CPT Code", "Procedure Code", "Service Code", "HCPCS"]
    notes_candidates = ["Code Notes", "Notes", "Additional Notes", "Comments", "Description", "MHI Code Notes"]
    med_candidates = ["Medicaid", "MHI Medicaid", "Medicaid PA", "Medicaid Status"]

    # Group mappings by out_name to support multiple source files mapping to one target file/sheet
    grouped_mappings = {}
    for m in mappings:
        out_s = m["out_name"]   # raw (long) name used as grouping key
        if out_s not in grouped_mappings:
            grouped_mappings[out_s] = {
                "sources": [],
                "source_aliases": [],
                "target_file": m["tgt_file"],
                "target_sheet": m["tgt_sheet"],
                "tgt_alias": m.get("tgt_alias", ""),
            }
        grouped_mappings[out_s]["sources"].append((m["src_file"], m["src_sheet"]))
        grouped_mappings[out_s]["source_aliases"].append(m.get("src_alias", ""))

    total = len(grouped_mappings)
    sanitized_to_raw = {}   # sheet_name (short) → original out_name (for metadata lookup)
    for i, (out_s_raw, config) in enumerate(grouped_mappings.items()):
        out_s = sanitize_sheet_name(out_s_raw)   # enforce Excel 31-char limit
        sanitized_to_raw[out_s] = out_s_raw
        if progress_callback: progress_callback(f"Processing {i+1}/{total}: {out_s_raw} (from {len(config['sources'])} sources)")

        try:
            # 1. Read Target File
            tgt_f = config["target_file"]
            tgt_s = config["target_sheet"]
            row_idx_new = find_header_row(tgt_f, tgt_s)
            df_new = pd.read_excel(tgt_f, sheet_name=tgt_s, header=row_idx_new, dtype=str)
            df_new = df_new.dropna(how='all', axis=1).dropna(how='all', axis=0)
            df_new = normalize_columns(df_new)

            c_code_new = find_column(df_new, code_candidates)
            c_notes_new = find_column(df_new, notes_candidates)
            c_med_new = find_column(df_new, med_candidates)
            if c_code_new: df_new.rename(columns={c_code_new: "STD_CODE"}, inplace=True)
            if c_notes_new: df_new.rename(columns={c_notes_new: "STD_NOTES"}, inplace=True)
            if c_med_new: df_new.rename(columns={c_med_new: "STD_MED"}, inplace=True)

            if not c_code_new:
                if progress_callback: progress_callback(f"⚠️ Skip {out_s}: Target ignores missing Code col.")
                continue

            # 2. Read and Combine Source Files
            # Tag each source row with its alias BEFORE concat so we can
            # assign quarter_id/year per code based on first-source priority
            source_aliases = config.get("source_aliases", [])
            source_dfs = []
            pad_aliases = source_aliases + [""] * len(config["sources"])
            for (src_f, src_s), src_alias in zip(config["sources"], pad_aliases):
                row_idx_old = find_header_row(src_f, src_s)
                df_old = pd.read_excel(src_f, sheet_name=src_s, header=row_idx_old, dtype=str)
                df_old = df_old.dropna(how='all', axis=1).dropna(how='all', axis=0)
                df_old = normalize_columns(df_old)

                c_code_old  = find_column(df_old, code_candidates)
                c_notes_old = find_column(df_old, notes_candidates)
                c_med_old   = find_column(df_old, med_candidates)

                if c_code_old:  df_old.rename(columns={c_code_old:  "STD_CODE"},  inplace=True)
                if c_notes_old: df_old.rename(columns={c_notes_old: "STD_NOTES"}, inplace=True)
                if c_med_old:   df_old.rename(columns={c_med_old:   "STD_MED"},   inplace=True)

                if "STD_CODE" in df_old.columns:
                    df_old["_source_alias"] = src_alias   # tag for quarter_id/year
                    source_dfs.append(df_old)

            if not source_dfs:
                if progress_callback: progress_callback(f"⚠️ Skip {out_s}: All sources missing Code col.")
                continue

            # Merge all sources (prioritizing later sources in dropping dupes inside compare_notes_and_medicaid logic)
            df_old_combined = pd.concat(source_dfs, ignore_index=True)

            notes_col_to_use = "STD_NOTES" if "STD_NOTES" in df_old_combined.columns and "STD_NOTES" in df_new.columns else None
            med_col_to_use = "STD_MED" if "STD_MED" in df_old_combined.columns and "STD_MED" in df_new.columns else None

            df_result = compare_notes_and_medicaid(
                df_old_combined, df_new,
                notes_col_to_use, med_col_to_use,
                source_aliases=source_aliases,
                tgt_alias=config.get("tgt_alias", "")
            )

            if df_result is not None and not df_result.empty:
                all_sheets_data[out_s] = df_result
            else:
                if progress_callback: progress_callback(f"⚠️ {out_s} produced no results.")
                
        except Exception as e:
            if progress_callback: progress_callback(f"❌ Error on {out_s}: {e}")

    if not all_sheets_data:
        if progress_callback: progress_callback("❌ No successful results generated.")
        return False

    if progress_callback: progress_callback("💾 Saving generated data to Excel...")
    
    with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
        for sheet_name, df in all_sheets_data.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    if progress_callback: progress_callback("🎨 Applying formatting...")
    try:
        wb = load_workbook(output_file)
        for sheet_name in all_sheets_data.keys():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                stats = {
                    "Total Source Codes": 0,
                    "Total Target Codes": 0,
                    "Changes (Modified)": 0,
                    "New Codes": 0,
                    "Removed Codes": 0
                }

                try:
                    headers = [cell.value for cell in ws[1]]
                    idx_status = headers.index("Status")
                    idx_severity = headers.index("Severity")

                    for row in ws.iter_rows(min_row=2, values_only=True):
                        st = row[idx_status]
                        if st != "New in Target":
                            stats["Total Source Codes"] += 1
                        if st != "Removed in Target":
                            stats["Total Target Codes"] += 1
                            
                        if st == "Modified": stats["Changes (Modified)"] += 1
                        if st == "New in Target": stats["New Codes"] += 1
                        if st == "Removed in Target": stats["Removed Codes"] += 1
                        
                except Exception:
                    pass

                write_summary(ws, stats,
                        source_aliases=grouped_mappings.get(sanitized_to_raw.get(sheet_name, sheet_name), {}).get("source_aliases"),
                        tgt_alias=grouped_mappings.get(sanitized_to_raw.get(sheet_name, sheet_name), {}).get("tgt_alias"))
                apply_conditional_formatting(ws)
                
        wb.save(output_file)
    except Exception as e:
        if progress_callback: progress_callback(f"⚠️ Formatting error: {e}")

    if progress_callback: progress_callback(f"✅ Success! Generated: {os.path.basename(output_file)}")
    return True

import os
import pandas as pd
import numpy as np
import google.generativeai as genai
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from dotenv import load_dotenv

# Load API key from .env
load_dotenv()
API_KEY = os.getenv("GOOGLE_API_KEY")
if not API_KEY:
    print("GOOGLE_API_KEY not found in .env")
else:
    API_KEY = 'AIzaSyD9WvnD_GR_rJcU2iZcjXlhtILRlbx5n_c'
genai.configure(api_key=API_KEY)

EMBED_MODEL = "models/text-embedding-004"


# ---------------------------
# Helpers
# ---------------------------
def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Strip whitespace/newlines from column names."""
    df.columns = df.columns.str.strip().str.replace('\n', ' ').str.replace('\r', ' ')
    return df


def find_column(df: pd.DataFrame, possible_names):
    """Find column name in df that matches one of possible_names (case-insensitive)."""
    lower_targets = [p.lower() for p in possible_names]
    for col in df.columns:
        if col.lower().strip() in lower_targets:
            return col
    raise KeyError(f"None of these columns found: {possible_names}. Available: {df.columns.tolist()}")


def embed(text: str):
    """Generate embedding vector using Gemini embeddings."""
    if text is None or str(text).strip() == "":
        text = "empty"
    resp = genai.embed_content(
        model=EMBED_MODEL,
        content=str(text),
        task_type="semantic_similarity"
    )
    return np.array(resp["embedding"])


def cosine_similarity(v1, v2):
    """Cosine similarity between two numpy vectors."""
    n1 = np.linalg.norm(v1)
    n2 = np.linalg.norm(v2)
    if n1 == 0 or n2 == 0:
        return 0.0
    return float(np.dot(v1, v2) / (n1 * n2))


def compute_summary(df: pd.DataFrame) -> dict:
    """Creates summary counts for the top of each sheet."""
    return {
        "Total in 2025 Q4": len(df),
        "No Change": int((df["Status"] == "No Change").sum()),
        "Modified": int((df["Status"] == "Modified").sum()),
        "Severe Change": int((df["Severity"].str.contains("Severe Change", na=False)).sum()),
        "Moderate Change": int((df["Severity"].str.contains("Moderate Change", na=False)).sum()),
        "Minor Change": int((df["Severity"].str.contains("Minor Wording Change", na=False)).sum()),
        "Medicaid Change": int((df["Severity"].str.contains("Medicaid Change", na=False)).sum()),
        "New in 2026 Q1": int((df["Status"] == "New in 2026 Q1").sum()),
        "Removed in 2026 Q1": int((df["Status"] == "Removed in 2026 Q1").sum())
    }


# ---------------------------
# Semantic + Medicaid comparison logic
# ---------------------------
def compare_notes_and_medicaid(q3_df: pd.DataFrame, q4_df: pd.DataFrame, notes_col: str, medicaid_col: str = None):
    """
    Compare two dataframes keyed on 'Code'.
    - If medicaid_col is provided: Compares Notes (Semantic) AND Medicaid (Exact string).
    - If medicaid_col is None: Compares Notes (Semantic) only.
    """
    # Ensure Code is string and drop duplicates by Code (keep first)
    q3_df = q3_df.copy()
    q4_df = q4_df.copy()

    # Although we read with dtype=str, we reinforce it here
    q3_df["Code"] = q3_df["Code"].astype(str).str.strip()
    q4_df["Code"] = q4_df["Code"].astype(str).str.strip()

    # Replace NaN with empty strings
    q3_df = q3_df.replace({np.nan: ""})
    q4_df = q4_df.replace({np.nan: ""})

    q4_lookup = {row["Code"]: row for _, row in q4_df.iterrows()}
    report_rows = []

    # Logic to handle if we are looking at medicaid column or not
    has_med_col = medicaid_col is not None

    for _, q3_row in q3_df.iterrows():
        code = q3_row["Code"]
        q3_notes = str(q3_row.get(notes_col, "")).strip()

        # Get Q3 Medicaid value if column exists, else ignore
        q3_med = str(q3_row.get(medicaid_col, "")).strip() if has_med_col else ""

        # Construct Q3 display string
        q3_val_str = f"Notes: {q3_notes}"
        if has_med_col:
            q3_val_str += f" | Medicaid: {q3_med}"

        # 1. Check if Code was removed
        if code not in q4_lookup:
            report_rows.append({
                "Code": code,
                "Status": "Removed in 2026 Q1",
                "Column": f"{notes_col}" + (f" / {medicaid_col}" if has_med_col else ""),
                "2025 Q4 Value": q3_val_str,
                "2026 Q1 Value": "",
                "Similarity": "",
                "Severity": "Severe Change"
            })
            continue

        q4_row = q4_lookup[code]
        q4_notes = str(q4_row.get(notes_col, "")).strip()
        q4_med = str(q4_row.get(medicaid_col, "")).strip() if has_med_col else ""

        # Construct Q4 display string
        q4_val_str = f"Notes: {q4_notes}"
        if has_med_col:
            q4_val_str += f" | Medicaid: {q4_med}"

        notes_same = (q3_notes == q4_notes)
        med_same = (q3_med == q4_med) if has_med_col else True

        # 2. Check Identical
        if notes_same and med_same:
            report_rows.append({
                "Code": code,
                "Status": "No Change",
                "Column": "",
                "2025 Q4 Value": q3_val_str,
                "2026 Q1 Value": q4_val_str,
                "Similarity": 1.0,
                "Severity": "No Change"
            })
            continue

        # 3. Check Changes
        severity_parts = []
        similarity_val = ""

        # Notes semantic comparison if changed
        if not notes_same:
            try:
                emb1 = embed(q3_notes)
                emb2 = embed(q4_notes)
                sim = cosine_similarity(emb1, emb2)
            except Exception:
                sim = 0.0
            similarity_val = round(sim, 4)

            if sim < 0.55:
                severity_parts.append("Severe Change")
            elif sim < 0.80:
                severity_parts.append("Moderate Change")
            else:
                severity_parts.append("Minor Wording Change")

        # Medicaid exact comparison
        if has_med_col and not med_same:
            severity_parts.insert(0, "Medicaid Change")

        severity_str = "; ".join(severity_parts) if severity_parts else "Modified"

        changed_cols = []
        if not notes_same:
            changed_cols.append(notes_col)
        if has_med_col and not med_same:
            changed_cols.append(medicaid_col)

        report_rows.append({
            "Code": code,
            "Status": "Modified",
            "Column": ", ".join(changed_cols),
            "2025 Q4 Value": q3_val_str,
            "2026 Q1 Value": q4_val_str,
            "Similarity": similarity_val,
            "Severity": severity_str
        })

    # 4. Check New Codes in Q4
    q3_codes = set(q3_df["Code"])
    q4_codes = set(q4_df["Code"])
    new_codes = q4_codes - q3_codes

    for code in new_codes:
        row = q4_df[q4_df["Code"] == code].iloc[0]
        q4_notes = str(row.get(notes_col, "")).strip()
        q4_med = str(row.get(medicaid_col, "")).strip() if has_med_col else ""

        q4_val_str = f"Notes: {q4_notes}"
        if has_med_col:
            q4_val_str += f" | Medicaid: {q4_med}"

        report_rows.append({
            "Code": code,
            "Status": "New in 2026 Q1",
            "Column": f"{notes_col}" + (f" / {medicaid_col}" if has_med_col else ""),
            "2025 Q4 Value": "",
            "2026 Q1 Value": q4_val_str,
            "Similarity": "",
            "Severity": "New Entry"
        })

    df_report = pd.DataFrame(report_rows, columns=[
        "Code", "Status", "Column", "2025 Q4 Value", "2026 Q1 Value", "Similarity", "Severity"
    ])
    return df_report.replace({np.nan: ""})


# ---------------------------
# Formatting Logic
# ---------------------------
def write_summary(ws, summary: dict):
    ws.insert_rows(1, amount=10)
    ws["A1"] = "SUMMARY"
    ws["A1"].font = Font(bold=True)
    row = 2
    for key, val in summary.items():
        ws[f"A{row}"] = key
        ws[f"B{row}"] = val
        row += 1


def apply_conditional_formatting(ws):
    colors = {
        "Severe Change": "FFC7CE",  # Red
        "Moderate Change": "FFEB9C",  # Yellow
        "Minor Wording Change": "C6EFCE",  # Green
        "New Entry": "BDD7EE",  # Blue
        "No Change": "FFFFFF",  # White
        "Medicaid Change": "E1D3F6"  # Purple-ish
    }
    for row in range(12, ws.max_row + 1):
        cell = ws.cell(row=row, column=7)
        severity_val = str(cell.value) if cell.value is not None else ""
        fill_color = None

        if "Medicaid Change" in severity_val:
            fill_color = colors["Medicaid Change"]
        elif "Severe Change" in severity_val:
            fill_color = colors["Severe Change"]
        elif "Moderate Change" in severity_val:
            fill_color = colors["Moderate Change"]
        elif "Minor Wording Change" in severity_val:
            fill_color = colors["Minor Wording Change"]
        elif "New Entry" in severity_val:
            fill_color = colors["New Entry"]
        elif "No Change" in severity_val:
            fill_color = colors["No Change"]

        if fill_color:
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill


# ---------------------------
# Main process
# ---------------------------
def process_file(excel_path, comparisons):
    """
    excel_path: path to the excel file
    comparisons: list of dicts, e.g.
    [
      {
        "q3_sheet": "WA Q3",
        "q4_sheet": "WA Q4",
        "output_sheet": "WA Report",
        "notes_col_candidates": ["Code Notes"],
        "medicaid_col_candidates": ["Medicaid"] (Optional)
      }
    ]
    """
    print(f"Reading Excel: {excel_path}...")
    xls = pd.ExcelFile(excel_path)

    # Helper to read and normalize
    def read_clean(sheet_name):
        # dtype=str is CRITICAL here. It ensures "00123" is read as string "00123", not int 123.
        df = pd.read_excel(xls, sheet_name, dtype=str)
        df = normalize_columns(df)
        df = df.replace({np.nan: ""})
        return df

    # Prepare writer to save reports
    # mode='a' appends to existing file, if_sheet_exists='replace' overwrites specific sheets
    with pd.ExcelWriter(excel_path, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer:

        processed_sheets = []

        for comp in comparisons:
            q3_name = comp["q3_sheet"]
            q4_name = comp["q4_sheet"]
            out_name = comp["output_sheet"]

            print(f"Processing: {q3_name} vs {q4_name} -> {out_name}")

            try:
                df_q3 = read_clean(q3_name)
                df_q4 = read_clean(q4_name)
            except ValueError as e:
                print(f"  Error reading sheets: {e}")
                continue

            # Find columns dynamically based on candidates
            notes_col = find_column(df_q3, comp["notes_col_candidates"])

            # Check if medicaid column comparison is requested
            med_col = None
            if "medicaid_col_candidates" in comp and comp["medicaid_col_candidates"]:
                try:
                    med_col = find_column(df_q3, comp["medicaid_col_candidates"])
                    print(f"  (Including Medicaid column: {med_col})")
                except KeyError:
                    print("  (Medicaid column candidates not found, skipping Medicaid comparison)")

            # Run comparison
            report_df = compare_notes_and_medicaid(df_q3, df_q4, notes_col, med_col)

            # Write to Excel
            report_df.to_excel(writer, sheet_name=out_name, index=False)
            processed_sheets.append(out_name)

    # Re-open with openpyxl for formatting (colors/summaries)
    print("Applying formatting...")
    wb = load_workbook(excel_path)

    for sheet_name in processed_sheets:
        ws = wb[sheet_name]

        # We need to re-calculate summary because we didn't return it from compare function
        # easier to just read the dataframe we just wrote or calculate from logic.
        # However, since the sheet is already written, let's just parse the dataframe from memory
        # But wait, we didn't save the report_df in memory in a persistent way.
        # Let's just re-read the dataframe from the writer context? No, that's closed.
        # Let's just read the sheet back into pandas? Or easier: modify compare to return summary?
        # Simplest: Just read the data from the openpyxl worksheet to calculate summary stats

        # Load data from worksheet to dict list
        data = ws.values
        cols = next(data)  # headers
        idx_status = -1
        idx_severity = -1

        # Find indices manually
        for i, c in enumerate(cols):
            if c == "Status": idx_status = i
            if c == "Severity": idx_severity = i

        # Tally up
        stats = {
            "Total in Q3": 0, "No Change": 0, "Modified": 0, "Severe Change": 0,
            "Moderate Change": 0, "Minor Wording Change": 0, "Medicaid Change": 0,
            "New in Q4": 0, "Removed in Q4": 0
        }

        for row in data:
            stats["Total in Q3"] += 1  # Approximation (includes new/removed) - actually total rows
            status_val = row[idx_status]
            severity_val = str(row[idx_severity])

            if status_val == "No Change": stats["No Change"] += 1
            if status_val == "Modified": stats["Modified"] += 1
            if status_val == "New in Q4": stats["New in Q4"] += 1
            if status_val == "Removed in Q4": stats["Removed in Q4"] += 1

            if "Severe Change" in severity_val: stats["Severe Change"] += 1
            if "Moderate Change" in severity_val: stats["Moderate Change"] += 1
            if "Minor Wording Change" in severity_val: stats["Minor Wording Change"] += 1
            if "Medicaid Change" in severity_val: stats["Medicaid Change"] += 1

        # Correct "Total in Q3" to be Total Rows - New in Q4? Or just Total Lines?
        # The original logic used len(df). Let's stick to total processed rows for simplicity in this dynamic version

        write_summary(ws, stats)
        apply_conditional_formatting(ws)

    wb.save(excel_path)
    print(f"\n🎉 Done! Updated: {excel_path}")


if __name__ == "__main__":
    file_path = "UNV 5 States.xlsx"

    # DEFINE YOUR DYNAMIC SHEETS HERE
    my_comparisons = [
        {
            "q3_sheet": "ID Q3",
            "q4_sheet": "ID Q4",
            "output_sheet": "ID Q3 vs ID Q4",
            "notes_col_candidates": ["Code Notes"],
            "medicaid_col_candidates": ["Medicaid"]
        },
        {
            "q3_sheet": "KY Q3",
            "q4_sheet": "KY Q4",
            "output_sheet": "KY Q3 vs KY Q4",
            "notes_col_candidates": ["Code Notes"],
            "medicaid_col_candidates": ["Medicaid"]
        },
        {
            "q3_sheet": "MS Q3",
            "q4_sheet": "MS Q4",
            "output_sheet": "MS Q3 vs MS Q4",
            "notes_col_candidates": ["Code Notes"],
            "medicaid_col_candidates": ["Medicaid"]
        },
        {
            "q3_sheet": "NY Q3",
            "q4_sheet": "NY Q4",
            "output_sheet": "NY Q3 vs NY Q4",
            "notes_col_candidates": ["Code Notes"],
            "medicaid_col_candidates": ["Medicaid"]
        },
        {
            "q3_sheet": "MA Q3",
            "q4_sheet": "MA Q4",
            "output_sheet": "MA Q3 vs MA Q4",
            "notes_col_candidates": ["Code Notes"],
            "medicaid_col_candidates": ["Medicaid"]
        },
    ]

    process_file(file_path, my_comparisons)
import os
import pandas as pd
import numpy as np
from difflib import SequenceMatcher
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font


def normalize_columns(df):
    df.columns = df.columns.str.strip().str.replace('\n', ' ').str.replace('\r', ' ')
    return df


def find_column(df, possible_names):
    for col in df.columns:
        if col.lower().strip() in [name.lower() for name in possible_names]:
            return col
    raise KeyError(f"❌ None of these columns found: {possible_names}")


def text_similarity(a, b):
    if not a and not b:
        return 1.0
    return SequenceMatcher(None, str(a), str(b)).ratio()


def compute_summary(df):
    return {
        "Total in Q3": len(df),
        "No Change": (df["Status"] == "No Change").sum(),
        "Modified": (df["Status"] == "Modified").sum(),
        "Severe Change": (df["Severity"] == "Severe Change").sum(),
        "Moderate Change": (df["Severity"] == "Moderate Change").sum(),
        "Minor Change": (df["Severity"] == "Minor Wording Change").sum(),
        "New in Q4": (df["Status"] == "New in Q4").sum(),
        "Removed in Q4": (df["Status"] == "Removed in Q4").sum()
    }


def semantic_compare(q3_df, q4_df, text_column):
    q3_df["Code"] = q3_df["Code"].astype(str)
    q4_df["Code"] = q4_df["Code"].astype(str)

    q4_lookup = {row["Code"]: row for _, row in q4_df.iterrows()}
    report_rows = []

    for _, q3_row in q3_df.iterrows():
        code = q3_row["Code"]

        if code not in q4_lookup:
            report_rows.append({
                "Code": code,
                "Status": "Removed in Q4",
                "Column": text_column,
                "Q3 Value": q3_row[text_column],
                "Q4 Value": "",
                "Similarity": "",
                "Severity": "Severe Change"
            })
            continue

        q4_row = q4_lookup[code]
        old = str(q3_row[text_column]).strip()
        new = str(q4_row[text_column]).strip()

        if old == new:
            similarity = 1.0
            severity = "No Change"
            status = "No Change"
        else:
            similarity = text_similarity(old, new)

            if similarity < 0.4:
                severity = "Severe Change"
            elif similarity < 0.75:
                severity = "Moderate Change"
            else:
                severity = "Minor Wording Change"

            status = "Modified"

        report_rows.append({
            "Code": code,
            "Status": status,
            "Column": text_column,
            "Q3 Value": old,
            "Q4 Value": new,
            "Similarity": round(similarity, 4),
            "Severity": severity
        })

    # Find new codes in Q4
    q3_codes = set(q3_df["Code"])
    q4_codes = set(q4_df["Code"])

    for code in q4_codes - q3_codes:
        q4_row = q4_df[q4_df["Code"] == code].iloc[0]
        report_rows.append({
            "Code": code,
            "Status": "New in Q4",
            "Column": text_column,
            "Q3 Value": "",
            "Q4 Value": q4_row[text_column],
            "Similarity": "",
            "Severity": "New Entry"
        })

    return pd.DataFrame(report_rows)


def apply_conditional_formatting(ws):
    colors = {
        "Severe Change": "FFC7CE",
        "Moderate Change": "FFEB9C",
        "Minor Wording Change": "C6EFCE",
        "New Entry": "BDD7EE",
        "No Change": "FFFFFF"
    }

    for row in range(12, ws.max_row + 1):
        severity = ws[f"G{row}"].value

        if severity in colors:
            fill = PatternFill(start_color=colors[severity], end_color=colors[severity], fill_type="solid")
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill


def write_summary(ws, summary):
    ws.insert_rows(1, amount=10)

    ws["A1"] = "SUMMARY"
    ws["A1"].font = Font(bold=True)

    row = 2
    for key, val in summary.items():
        ws[f"A{row}"] = key
        ws[f"B{row}"] = val
        row += 1


def process_file(excel_path):
    print("Reading Excel...")
    xls = pd.ExcelFile(excel_path)

    wa_q3 = normalize_columns(pd.read_excel(xls, "WA Q3"))
    wa_q4 = normalize_columns(pd.read_excel(xls, "WA Q4"))
    md_q3 = normalize_columns(pd.read_excel(xls, "Medicaid Q3"))
    md_q4 = normalize_columns(pd.read_excel(xls, "Medicaid Q4"))

    wa_q3 = wa_q3.replace(np.nan, "", regex=True)
    wa_q4 = wa_q4.replace(np.nan, "", regex=True)
    md_q3 = md_q3.replace(np.nan, "", regex=True)
    md_q4 = md_q4.replace(np.nan, "", regex=True)

    wa_col_q3 = find_column(wa_q3, ["Code Notes"])
    wa_report = semantic_compare(wa_q3, wa_q4, wa_col_q3)

    md_col_q3 = find_column(md_q3, ["MHI Code Notes"])
    md_report = semantic_compare(md_q3, md_q4, md_col_q3)

    wa_report = wa_report.replace(np.nan, "")
    md_report = md_report.replace(np.nan, "")

    with pd.ExcelWriter(excel_path, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer:
        wa_report.to_excel(writer, sheet_name="WA Q3 vs WA Q4", index=False)
        md_report.to_excel(writer, sheet_name="Medicaid Q3 vs Medicaid Q4", index=False)

    wb = load_workbook(excel_path)

    ws_wa = wb["WA Q3 vs WA Q4"]
    ws_md = wb["Medicaid Q3 vs Medicaid Q4"]

    write_summary(ws_wa, compute_summary(wa_report))
    write_summary(ws_md, compute_summary(md_report))

    apply_conditional_formatting(ws_wa)
    apply_conditional_formatting(ws_md)

    wb.save(excel_path)

    print("\n🎉 Done!")
    print("Sheets updated:")
    print("→ WA Q3 vs WA Q4")
    print("→ Medicaid Q3 vs Medicaid Q4")
    print("✔ Local-only comparison")
    print("✔ Color-coded output")
    print("✔ Summary added")


if __name__ == "__main__":
    process_file("Authorization Business Matrix 2025 Q3 - WA and Medicaid - Reference.xlsx")

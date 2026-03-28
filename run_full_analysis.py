import pandas as pd
import os
import sys
import warnings
from openpyxl import load_workbook

# Suppress warnings
warnings.filterwarnings("ignore")


def find_header_row(xls, sheet_name):
    """
    Scans the first 50 rows to find the true header row by looking for 'Code' column.
    Returns: (index, found_boolean)
    """
    try:
        df_preview = pd.read_excel(xls, sheet_name=sheet_name, header=None, nrows=50, dtype=str)
    except Exception:
        return 0, False

    target_headers = ["code", "cpt code", "procedure code", "hcpcs", "service code"]

    for idx, row in df_preview.iterrows():
        row_values = [str(x).strip().lower() for x in row.values]
        if any(t in row_values for t in target_headers):
            return idx, True

    return 0, False


def main():
    print("=" * 60)
    print("       🚀 OPTIMIZED ANALYSIS PIPELINE (V5 - ROBUST)")
    print("=" * 60)

    # 1. CONFIGURATION
    file_old = "Authorization Business Matrix 2025 Q4 - All States and LOBs - Reference.xlsx"
    file_new = "Authorization Business Matrix 2026 Q1 - All States and LOBs - Reference.xlsx"
    final_output_file = "Final_2025Q4_vs_2026Q1_Analysis.xlsx"

    # 2. CHECK FILES
    if not os.path.exists(file_old):
        print(f"❌ Error: File not found: {file_old}")
        return
    if not os.path.exists(file_new):
        print(f"❌ Error: File not found: {file_new}")
        return
    if not os.path.exists("new_code_analysis.py"):
        print("❌ Error: 'new_code_analysis.py' not found.")
        return

    # 3. IMPORT ANALYSIS LOGIC
    print("🔄 Importing analysis logic...")
    try:
        import new_code_analysis
    except ImportError as e:
        print(f"❌ Error importing script: {e}")
        return

    # 4. LOAD MASTER FILES
    print("\n⏳ Loading Master Excel Files into Memory...")
    try:
        xls_old = pd.ExcelFile(file_old)
        print(f"   Reading 2025 Q4: {file_old}")
        xls_new = pd.ExcelFile(file_new)
        print(f"   Reading 2026 Q1: {file_new}")
    except Exception as e:
        print(f"❌ Error reading Excel files: {e}")
        return

    # 5. IDENTIFY COMMON SHEETS
    ignored_sheets = [
        "Sheet1", "UPDATES", "Updates", "UPDATES Temp", "Lookup Tool",
        "ICD 10 Codes", "Evolent Delegated Codes", "MEDICARE", "MEDICAID",
        "MARKETPLACE", "MEDICAID Evolent Archive", "Reference",
        "Change Log", "Instructions", "Table of Contents", "Introduction"
    ]

    sheets_old = set(xls_old.sheet_names)
    sheets_new = set(xls_new.sheet_names)
    common_sheets = sorted([s for s in sheets_old if s in sheets_new and s not in ignored_sheets])

    if not common_sheets:
        print("❌ No matching state sheets found.")
        return

    print(f"✅ Found {len(common_sheets)} states to process.")

    # 6. PROCESS ALL STATES IN MEMORY
    print("\n" + "-" * 30)
    print("⚡ PROCESSING & COMPARING STATES")
    print("-" * 30)

    all_sheets_data = {}
    analysis_sheet_names = []

    code_candidates = ["Code", "CPT Code", "Procedure Code", "Service Code", "HCPCS"]
    notes_candidates = ["Code Notes", "Notes", "Additional Notes", "Comments", "Description"]
    med_candidates = ["Medicaid", "MHI Medicaid", "Medicaid PA", "Medicaid Status"]

    for sheet in common_sheets:
        try:
            # HEADER DETECTION
            row_idx_old, found_old = find_header_row(xls_old, sheet)
            row_idx_new, found_new = find_header_row(xls_new, sheet)

            status_msg = f"   ... Processing {sheet:<5} "
            status_msg += f"[Headers: {row_idx_old}/{row_idx_new}]" if (found_old or found_new) else "[Default Header]"
            print(status_msg)

            # Read DataFrames
            df_old = pd.read_excel(xls_old, sheet_name=sheet, header=row_idx_old, dtype=str)
            df_new = pd.read_excel(xls_new, sheet_name=sheet, header=row_idx_new, dtype=str)

            # Drop empty columns/rows
            df_old = df_old.dropna(how='all', axis=1).dropna(how='all', axis=0)
            df_new = df_new.dropna(how='all', axis=1).dropna(how='all', axis=0)

            # Normalize Headers
            df_old = new_code_analysis.normalize_columns(df_old)
            df_new = new_code_analysis.normalize_columns(df_new)

            # CHECK FOR NONE BEFORE STORING
            if df_old is None or df_new is None:
                print(f"       ⚠️ SKIP: Failed to read data correctly for {sheet}")
                continue

            # VALIDATION
            col_old = new_code_analysis.find_column(df_old, code_candidates)
            col_new = new_code_analysis.find_column(df_new, code_candidates)

            if not col_old:
                print(f"       ⚠️ SKIP: 'Code' column missing in 2025 file.")
                continue
            if not col_new:
                print(f"       ⚠️ SKIP: 'Code' column missing in 2026 file.")
                continue

            name_old = f"{sheet} 25Q4"
            name_new = f"{sheet} 26Q1"
            name_analysis = f"{sheet} Comparison"

            # Store Raw Data (Guaranteed valid DFs now)
            all_sheets_data[name_old] = df_old
            all_sheets_data[name_new] = df_new

            # Find columns for analysis
            notes_col = new_code_analysis.find_column(df_old, notes_candidates)
            med_col = new_code_analysis.find_column(df_old, med_candidates)

            # Run Comparison
            df_result = new_code_analysis.compare_notes_and_medicaid(df_old, df_new, notes_col, med_col)

            # DEFENSIVE CHECK: Ensure result is valid
            if df_result is not None and not df_result.empty:
                all_sheets_data[name_analysis] = df_result
                analysis_sheet_names.append(name_analysis)
            else:
                print(f"       ⚠️ WARNING: Analysis returned no results for {sheet}.")

        except Exception as e:
            print(f"       ❌ Error: {e}")

    # 7. WRITE EVERYTHING TO EXCEL
    if not all_sheets_data:
        print("\n❌ No data was processed. Exiting.")
        return

    print("\n💾 Writing all data to Excel (Using fast xlsxwriter engine)...")

    with pd.ExcelWriter(final_output_file, engine='xlsxwriter') as writer:
        for sheet_name, df in all_sheets_data.items():
            # FINAL SAFETY CHECK
            if df is None:
                print(f"       ⚠️ Skipping write for {sheet_name} (Data is None)")
                continue

            try:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            except Exception as e:
                print(f"       ❌ Failed to write sheet {sheet_name}: {e}")

    print(f"✅ Saved raw data to {final_output_file}")

    # 8. APPLY FORMATTING
    print("🎨 Applying conditional formatting (Colors)...")

    try:
        wb = load_workbook(final_output_file)

        for sheet_name in analysis_sheet_names:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Calculate summary stats manually
                stats = {
                    "Total in 2025 Q4": 0, "No Change": 0, "Modified": 0, "Severe Change": 0,
                    "Moderate Change": 0, "Minor Wording Change": 0, "Medicaid Change": 0,
                    "New in 2026 Q1": 0, "Removed in 2026 Q1": 0
                }

                headers = [cell.value for cell in ws[1]]
                try:
                    idx_status = headers.index("Status")
                    idx_severity = headers.index("Severity")

                    for row in ws.iter_rows(min_row=2, values_only=True):
                        stats["Total in 2025 Q4"] += 1
                        status_val = row[idx_status]
                        severity_val = str(row[idx_severity]) if row[idx_severity] else ""

                        if status_val == "No Change": stats["No Change"] += 1
                        if status_val == "Modified": stats["Modified"] += 1
                        if status_val == "New in 2026 Q1": stats["New in 2026 Q1"] += 1
                        if status_val == "Removed in 2026 Q1": stats["Removed in 2026 Q1"] += 1

                        if "Severe Change" in severity_val: stats["Severe Change"] += 1
                        if "Moderate Change" in severity_val: stats["Moderate Change"] += 1
                        if "Minor Wording Change" in severity_val: stats["Minor Wording Change"] += 1
                        if "Medicaid Change" in severity_val: stats["Medicaid Change"] += 1
                except ValueError:
                    pass

                new_code_analysis.write_summary(ws, stats)
                new_code_analysis.apply_conditional_formatting(ws)

        wb.save(final_output_file)
        print("✅ Formatting complete.")

    except Exception as e:
        print(f"⚠️ Warning: Could not apply formatting: {e}")

    print("\n" + "=" * 60)
    print(f"🎉 DONE! Final Output: {final_output_file}")
    print("=" * 60)


if __name__ == "__main__":
    main()
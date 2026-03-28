import ast
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ─────────────────────────────────────────────
# CONSTANTS  –  edit these before running
# ─────────────────────────────────────────────

# List the sheet names you want the script to process.
# Leave empty  →  process ALL sheets in the workbook.
TARGET_SHEETS: list[str] = [
    "MS",
    "ID",
    "NY",
]

# Value to place in every data cell of the "api response" column.
# Allowed values:  ""  (empty)  |  "Match"  |  "No Match"
MATCH_VALUE: str = ""

# Input file path
INPUT_FILE: str = "/Users/xaiuser/Downloads/Copy of batch_report_UNV_367_2026-03.xlsx"

# Directory where per-sheet output files are saved (use "." for current folder)
OUTPUT_DIR: str = "."

# DENY / OKAY status count thresholds
STATUS_MIN: int = 10   # minimum expected count  (triggers white-row fallback if below)
STATUS_MAX: int = 15   # maximum allowed  count

# ─────────────────────────────────────────────
# Internal column name constants
# ─────────────────────────────────────────────
API_RESPONSE_COL   = "api response"
JUSTIFICATION_COL  = "justification"
RECOMMENDATION_COL = "recommendation"
STATE_COL          = "state"
YELLOW_HEX         = "FFFF00"

# ─────────────────────────────────────────────


# ── Recommendation helpers ───────────────────────────────────────────────────

def parse_recommendation(value) -> dict | None:
    """Parse a recommendation column cell (Python dict string) into a dict."""
    if pd.isna(value) or not str(value).strip():
        return None
    try:
        return ast.literal_eval(str(value))
    except Exception:
        return None


def get_status_counts(df: pd.DataFrame) -> dict[str, int]:
    """
    Parse the recommendation column and return a dict of status → count.
    Only DENY and OKAY are tracked.
    """
    counts: dict[str, int] = {"DENY": 0, "OKAY": 0}
    if RECOMMENDATION_COL not in df.columns:
        return counts
    for cell_value in df[RECOMMENDATION_COL]:
        parsed = parse_recommendation(cell_value)
        if not parsed:
            continue
        for rec in parsed.get("recommendation", []):
            status = str(rec.get("status", "")).upper()
            if status in counts:
                counts[status] += 1
    return counts


def print_status_validation(counts: dict[str, int], sheet_name: str) -> None:
    """Print DENY / OKAY counts, valid when STATUS_MIN <= count <= STATUS_MAX."""
    print(f"\n  [{sheet_name}] 📊 Status counts (from '{RECOMMENDATION_COL}' column):")
    for status in ("DENY", "OKAY"):
        count = counts.get(status, 0)
        valid = STATUS_MIN <= count <= STATUS_MAX
        flag = "✅" if valid else "❌"
        print(
            f"    {flag}  {status}: {count}  "
            f"(expected {STATUS_MIN}–{STATUS_MAX}, "
            f"{'VALID' if valid else 'OUT OF RANGE'})"
        )


# ── Fill helpers ─────────────────────────────────────────────────────────────

def get_yellow_row_indices(wb_path: str, sheet_name: str) -> set[int] | None:
    """
    Return 0-based DataFrame row indices that have a yellow background fill.
    Returns None if the sheet is not found.
    """
    wb = load_workbook(wb_path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return None

    ws = wb[sheet_name]
    yellow_indices: set[int] = set()

    for openpyxl_row in ws.iter_rows(min_row=2):   # skip header
        for cell in openpyxl_row:
            fill = cell.fill
            if fill and fill.fill_type == "solid":
                fg = fill.fgColor
                if fg and fg.type == "rgb" and fg.rgb.upper().endswith("FFFF00"):
                    yellow_indices.add(cell.row - 2)   # openpyxl row → df index
                    break

    wb.close()
    return yellow_indices


def apply_yellow_fill(filepath: str) -> None:
    """Paint every cell in every sheet of the given workbook yellow."""
    wb = load_workbook(filepath)
    yellow_fill = PatternFill(
        start_color=YELLOW_HEX, end_color=YELLOW_HEX, fill_type="solid"
    )
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.fill = yellow_fill
    wb.save(filepath)


# ── Core processing ──────────────────────────────────────────────────────────

def split_and_save(df: pd.DataFrame, sheet_name: str, source: str) -> None:
    """
    Split df into Marketplace / Medicaid subsets and save each as a separate file.
    - source: a label like "yellow" or "white" used only for console output.
    """
    def save_subset(df_sub: pd.DataFrame, label: str) -> None:
        if df_sub.empty:
            print(f"           ⏭️   No '{label}' rows ({source}) — file skipped.")
            return

        # Cap this subset at STATUS_MAX rows
        if len(df_sub) > STATUS_MAX:
            df_sub = df_sub.head(STATUS_MAX).copy()
            print(f"           ✂️   [{label}] Capped to {STATUS_MAX} rows.")

        # Clear / set the api response column
        if API_RESPONSE_COL in df_sub.columns:
            df_sub = df_sub.copy()
            df_sub[API_RESPONSE_COL] = MATCH_VALUE

        output_file = os.path.join(OUTPUT_DIR, f"{sheet_name}_Validation_{label}.xlsx")
        with pd.ExcelWriter(output_file, engine="openpyxl", mode="w") as writer:
            df_sub.to_excel(writer, sheet_name=sheet_name, index=False)
        apply_yellow_fill(output_file)

        counts = get_status_counts(df_sub)
        print_status_validation(counts, sheet_name)
        print(f"           ✅  [{label}] Saved → '{output_file}'  ({len(df_sub)} rows, source: {source})")

    # ── Split ────────────────────────────────────────────────────────────────
    if JUSTIFICATION_COL in df.columns:
        mask = df[JUSTIFICATION_COL].astype(str).str.startswith("Marketplace", na=False)
        df_marketplace = df[mask].copy()
        df_medicaid    = df[~mask].copy()
    else:
        print(f"  [{sheet_name}] ⚠️   '{JUSTIFICATION_COL}' column not found — all rows → Medicaid.")
        df_marketplace = pd.DataFrame()
        df_medicaid    = df.copy()

    print(
        f"  [{sheet_name}] 📂  Split ({source}) → "
        f"Marketplace: {len(df_marketplace)} rows | Medicaid: {len(df_medicaid)} rows"
    )
    save_subset(df_marketplace, "Marketplace")
    save_subset(df_medicaid,    "Medicaid")


def filter_by_state(df: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
    """Keep only rows where the state column matches sheet_name."""
    if STATE_COL not in df.columns:
        return df
    return df[
        df[STATE_COL].astype(str).str.strip().str.upper() == sheet_name.upper()
    ].copy()


# ── Main ─────────────────────────────────────────────────────────────────────

def process_workbook(input_file: str) -> None:
    # 1. Load all sheets via pandas
    try:
        all_sheets: dict[str, pd.DataFrame] = pd.read_excel(
            input_file, sheet_name=None, dtype=str
        )
    except FileNotFoundError:
        print(f"❌  Error: '{input_file}' not found.")
        return

    sheets_to_process = list(all_sheets.keys()) if not TARGET_SHEETS else TARGET_SHEETS
    missing = [s for s in sheets_to_process if s not in all_sheets]
    if missing:
        print(f"⚠️   Sheets not found (skipped): {missing}")
    sheets_to_process = [s for s in sheets_to_process if s in all_sheets]

    print(f"\n📂  Input : {input_file}")
    print(f"📁  Output: {os.path.abspath(OUTPUT_DIR)}\n")

    for sheet_name in sheets_to_process:
        print(f"\n  ── {sheet_name} ──────────────────────────────────────")
        df_all = all_sheets[sheet_name].copy()

        # ── a. Get yellow row indices ─────────────────────────────────────
        yellow_indices = get_yellow_row_indices(input_file, sheet_name)
        if yellow_indices is None:
            print(f"  [{sheet_name}] ⚠️   Sheet not found in workbook — skipping.")
            continue

        # ── b. Yellow rows → filter by state → filter by Match ───────────
        df_yellow = df_all[df_all.index.isin(yellow_indices)].copy()
        print(f"  [{sheet_name}] 🟡  Yellow rows: {len(df_all)} total → {len(df_yellow)} yellow")

        df_yellow = filter_by_state(df_yellow, sheet_name)
        print(f"  [{sheet_name}] 🔍  State filter: → {len(df_yellow)} rows")

        match_mask = df_yellow.apply(
            lambda row: row.astype(str).str.strip().str.lower().eq("match").any(),
            axis=1,
        )
        df_match = df_yellow[match_mask].copy()
        print(f"  [{sheet_name}] ✔️   Any cell == 'Match': → {len(df_match)} rows")

        # ── c. Split yellow+match rows by justification ──────────────────
        if JUSTIFICATION_COL in df_match.columns:
            mp_mask  = df_match[JUSTIFICATION_COL].astype(str).str.startswith("Marketplace", na=False)
            df_match_mp = df_match[mp_mask].copy()
            df_match_md = df_match[~mp_mask].copy()
        else:
            df_match_mp = pd.DataFrame()
            df_match_md = df_match.copy()

        # ── d. Get white rows (non-yellow, state-filtered) and split them ─
        df_white = df_all[~df_all.index.isin(yellow_indices)].copy()
        df_white  = filter_by_state(df_white, sheet_name)
        print(f"  [{sheet_name}] ⬜  White rows available: {len(df_white)}")

        if JUSTIFICATION_COL in df_white.columns:
            wp_mask  = df_white[JUSTIFICATION_COL].astype(str).str.startswith("Marketplace", na=False)
            df_white_mp = df_white[wp_mask].copy()
            df_white_md = df_white[~wp_mask].copy()
        else:
            df_white_mp = pd.DataFrame()
            df_white_md = df_white.copy()

        # ── e. Build each subset independently ───────────────────────────
        def build_subset(df_ym: pd.DataFrame, df_w: pd.DataFrame, label: str) -> pd.DataFrame:
            """
            If yellow+match rows are ≥ STATUS_MIN → use only yellow+match (cap at 15).
            Else → supplement with white rows of the same type up to STATUS_MAX (15).
            """
            if len(df_ym) >= STATUS_MIN:
                result = df_ym.head(STATUS_MAX).copy()
                print(
                    f"  [{sheet_name}] ✅  [{label}] yellow+match ({len(df_ym)}) ≥ {STATUS_MIN}"
                    f" — using yellow+match only"
                )
            else:
                needed = max(0, STATUS_MAX - len(df_ym))
                df_w_capped = df_w.head(needed)
                result = pd.concat([df_ym, df_w_capped], ignore_index=True).drop_duplicates()
                result = result.head(STATUS_MAX)
                print(
                    f"  [{sheet_name}] 🔄  [{label}] yellow+match ({len(df_ym)}) < {STATUS_MIN}"
                    f" — added {len(df_w_capped)} white rows → {len(result)} total"
                )
            return result

        df_mp_final = build_subset(df_match_mp, df_white_mp, "Marketplace")
        df_md_final = build_subset(df_match_md, df_white_md, "Medicaid")

        print(f"  [{sheet_name}] 📂  Final → Marketplace: {len(df_mp_final)} rows | Medicaid: {len(df_md_final)} rows")

        # ── f. Save each subset ───────────────────────────────────────────
        def save_final(df_sub: pd.DataFrame, label: str) -> None:
            if df_sub.empty:
                print(f"           ⏭️   No '{label}' rows — file skipped.")
                return
            if API_RESPONSE_COL in df_sub.columns:
                df_sub = df_sub.copy()
                df_sub[API_RESPONSE_COL] = MATCH_VALUE
            output_file = os.path.join(OUTPUT_DIR, f"{sheet_name}_Validation_{label}.xlsx")
            with pd.ExcelWriter(output_file, engine="openpyxl", mode="w") as writer:
                df_sub.to_excel(writer, sheet_name=sheet_name, index=False)
            apply_yellow_fill(output_file)
            counts = get_status_counts(df_sub)
            print_status_validation(counts, sheet_name)
            print(f"           ✅  [{label}] Saved → '{output_file}'  ({len(df_sub)} rows)")

        save_final(df_mp_final, "Marketplace")
        save_final(df_md_final, "Medicaid")

    print("\n🎉  All done!")



# ─────────────────────────────────────────────
if __name__ == "__main__":
    process_workbook(INPUT_FILE)